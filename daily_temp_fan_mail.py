from itertools import count
import pandas as pd
import numpy as np
import glob
import os
import re
import win32com.client
import win32timezone
import os
from datetime import datetime, timedelta
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side

#================================================================================
# ✅ Fetching Last 7 Days' Reports from Outlook
#================================================================================

# # Define the folder where attachments will be saved
# save_path = r"D:\Daily_High_Temp_Alarm_Dump"
# os.makedirs(save_path, exist_ok=True)  # Ensure save directory exists

# # Connect to Outlook
# try:
#     outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
# except Exception as e:
#     print(f"❌ Error connecting to Outlook: {e}")
#     sys.exit()

# # Select the Inbox → Required to Work folder
# try:
#     inbox = outlook.GetDefaultFolder(6)  # Inbox folder (6)
#     target_folder = inbox.Folders("Required to Work")  # Navigate to subfolder
# except Exception as e:
#     print("❌ Error: 'Required to Work' folder not found!")
#     sys.exit()

# # Get all emails and sort by received time (latest first)
# messages = target_folder.Items
# messages.Sort("[ReceivedTime]", True)

# # Filter emails received in the last 3 days
# seven_days_ago = (datetime.now() - timedelta(days=3)).strftime('%m/%d/%Y')
# messages = messages.Restrict(f"[ReceivedTime] >= '{seven_days_ago}'")

# # Check if any emails exist
# if messages.Count == 0:
#     print("❌ No emails received in the last 3 days. Exiting!")
#     exit()
# else:
#     found = False
#     for message in messages:
#         try:
#             subject = message.Subject
#             received_time = message.ReceivedTime
#             print(f"📧 Checking email: {subject} | Date: {received_time.strftime('%Y-%m-%d')}")

#             if "RAN Daily FAN & Temp Alarms Details" in subject:
#                 print("✅ Found matching email!")

#                 if message.Attachments.Count > 0:
#                     for attachment in message.Attachments:
#                         if attachment.FileName.endswith(".xlsx"):
#                             mail_date = received_time.strftime('%Y%m%d')
#                             base_name, ext = os.path.splitext(attachment.FileName)
#                             new_file_name = f"{base_name}_{mail_date}{ext}"
#                             file_path = os.path.join(save_path, new_file_name)

#                             try:
#                                 attachment.SaveAsFile(file_path)
#                                 print(f"📂 Attachment saved: {file_path}")
#                                 found = True
#                             except Exception as e:
#                                 print(f"❌ Error saving attachment: {e}")
#                 else:
#                     print("⚠️ Email has no attachments.")

#         except Exception as e:
#             print(f"⚠️ Error processing email: {e}")

#     if not found:
#         print("❌ No email with 'RAN Daily FAN & Temp Alarms Details' found in the last 3 days.")

# # Release COM objects properly
# del messages, target_folder, inbox, outlook

# # ✅ Proceed with the next steps after downloading the attachments
# print("🚀 Proceeding to the next script execution...")


#=============================================================================Reading High Temp Fan Alarm Files============================================================
df_amob_acoc = pd.read_excel("D:/Nokia Cabinet Data/Nokia Cabinet Data.xlsx",sheet_name='Sheet1')

# Folder path containing the files
folder_path = "D:/Daily_High_Temp_Alarm_Dump"


# Get all .xlsx files and sort by filename (reverse for latest files first)
file_paths = glob.glob(os.path.join(folder_path, "*.xlsx"))
file_paths.sort(reverse=True)  # Alphabetically descending (latest file names first)



# Get only the latest 9 files
latest_files = file_paths[:17]

# Columns to load from each file
columns_to_load = ['Row Number', 'Circle', 'SITEID', 'Alarm Name', 'Node', 'Technology', 
                   'EMS Start Time', 'EMS End Time', 'Status', 'Alert Key', 'Additional Info', 'Agent']

# Initialize a list to hold data from each file
all_data = []

# Read each file and append the data to the list
for file_path in latest_files:
    try:
        file_size = os.path.getsize(file_path) / (1024 * 1024)  # Convert bytes to MB
        print(f"Reading file: {os.path.basename(file_path)} | Size: {file_size:.2f} MB")
        
        data = pd.read_excel(file_path, usecols=columns_to_load)
        all_data.append(data)
    except Exception as e:
        print(f"Error reading {file_path}: {e}")

# Concatenate all data into a single DataFrame
if all_data:
    combined_data_high_temp = pd.concat(all_data, ignore_index=True)
    print("Data successfully combined.")
else:
    combined_data_high_temp = pd.DataFrame(columns=columns_to_load)
    print("No valid data found.")


#===========================================================================processing combined data========================================================================

combined_data_high_temp = combined_data_high_temp.dropna(subset=['EMS Start Time'])
combined_data_high_temp.rename(columns={'Alarm Name': 'Alarm_Name','Additional Info': 'Additional_Info','Alert Key': 'Alert_Key'}, inplace=True)
#combined_data_high_temp = combined_data_high_temp.dropna(subset=['EMS Start Time'])
combined_data_high_temp['EMS Start Date'] = combined_data_high_temp['EMS Start Time'].dt.date
combined_data_high_temp['EMS Start Date'] = pd.to_datetime(combined_data_high_temp['EMS Start Date'], errors='coerce')
# Get today's date
today_date = pd.to_datetime('today').normalize()
# Filter out rows where 'EMS Start Date' is today
combined_data_high_temp = combined_data_high_temp[combined_data_high_temp['EMS Start Date'] < today_date]
combined_data_high_temp = combined_data_high_temp[combined_data_high_temp['SITEID'] != "Unknown SITEID"]
# Extract week number from 'EMS Start Date'
combined_data_high_temp['Week Number'] = combined_data_high_temp['EMS Start Date'].dt.isocalendar().week
combined_data_high_temp['Week Name'] = 'WK-' + combined_data_high_temp['EMS Start Date'].dt.isocalendar().week.astype(str)

# Extract unique dates and drop NaT values
unique_dates = combined_data_high_temp['EMS Start Date'].unique()
# Sort the unique dates and format them
readable_dates = [date.strftime('%d-%m-%Y') for date in sorted(unique_dates)]
# Print the unique, readable dates
print("Unique EMS Start Dates:")
for date in readable_dates:
    print(date)
try:
    # # Convert start date to datetime
    #start_date = "01-01-2025"
    #start_date = pd.to_datetime(start_date, format='%d-%m-%Y', errors='coerce')

    # Ensure 'EMS Start Date' is in datetime format
    combined_data_high_temp['EMS Start Date'] = pd.to_datetime(combined_data_high_temp['EMS Start Date'], errors='coerce')

    # Get max EMS Start Date
    end_date = combined_data_high_temp['EMS Start Date'].max()
    start_date = end_date - pd.Timedelta(days=15)

    # Check if dates are valid
    if pd.isna(start_date) or pd.isna(end_date):
        print("Invalid date format or missing data.")
    elif start_date > end_date:
        print("Start date cannot be later than the max EMS Start Date.")
    else:
        # Filter data
        filtered_data = combined_data_high_temp[
            (combined_data_high_temp['EMS Start Date'] >= start_date) &
            (combined_data_high_temp['EMS Start Date'] <= end_date)
        ]

        print(f"Filtered data between {start_date.strftime('%d-%m-%Y')} and {end_date.strftime('%d-%m-%Y')}:")

except Exception as e:
    print(f"An error occurred: {e}")
#======================================================================================================Working on OEM Column==================================================================================================
#Creating new column OEM
filtered_data = filtered_data.copy() 
filtered_data.loc[:,'OEM'] = filtered_data['Agent'].apply(
    lambda x: 'Ericsson' if 'Ericsson' in x else
              'Nokia' if 'Nokia' in x else
              'Huawei' if 'Huawei' in x else
              'ZTE' if 'ZTE' in x else
              'SAMSUNG' if 'SAMSUNG' in x else
              'IPA' if 'IPA' in x else
              'Unknown'
)

#=============================================================================================Working on FAN/Temperature alarm Column=========================================================================================

def classify_alarm_type(Additional_Info):
    if isinstance(Additional_Info, str):
        if 'Fan' in Additional_Info:
            return 'FAN alarm'
        elif 'fan' in Additional_Info:
            return 'FAN alarm'
        elif 'Temperature' in Additional_Info:
            return 'Temperature Alarm'
        elif 'temperature' in Additional_Info:
            return 'Temperature Alarm'
        elif 'overheating' in Additional_Info:
            return 'Temperature Alarm'
        elif 'Temp Fault' in Additional_Info:
            return 'Temp Fault'
        else:
            return np.nan
    return np.nan



filtered_data = filtered_data.copy() 
filtered_data.loc[filtered_data['OEM'] == "Nokia", 'FAN/Temperature alarm'] = filtered_data['Additional_Info'].apply(classify_alarm_type)
filtered_data = filtered_data[filtered_data['FAN/Temperature alarm'] != 'Temp Fault']



def classify_alarm_type(Additional_Info):
    if isinstance(Additional_Info, str):
        if 'FAN' in Additional_Info:
            return 'FAN alarm'
        elif 'temperature' in Additional_Info:
            return 'Temperature Alarm'        
        else:
            return np.nan
    return np.nan



filtered_data = filtered_data.copy() 
filtered_data.loc[filtered_data['OEM'] == "SAMSUNG", 'FAN/Temperature alarm'] = filtered_data['Additional_Info'].apply(classify_alarm_type)


def classify_alarm_type(Alarm_Name):
    if isinstance(Alarm_Name, str):
        if 'Fan' in Alarm_Name:
            return 'FAN alarm'
        elif 'fan' in Alarm_Name:
            return 'FAN alarm'
        elif 'FAN' in Alarm_Name:
            return 'FAN alarm'        
        elif 'TEMPERATURE' in Alarm_Name:
            return 'Temperature Alarm'
        elif 'Temperature' in Alarm_Name:
            return 'Temperature Alarm'
        elif 'temperature' in Alarm_Name:
            return 'Temperature Alarm'
        elif 'overtemperature' in Alarm_Name:
            return 'Temperature Alarm'
        else:
            return np.nan
    return np.nan

filtered_data.loc[filtered_data['OEM'].isin(["Ericsson","Huawei","ZTE"]), 'FAN/Temperature alarm'] = filtered_data['Alarm_Name'].apply(classify_alarm_type)

#=======================================================================================working on RRU/Baseband===============================================================================================================


def classify_rru_baseband(Additional_Info):
    if isinstance(Additional_Info, str):
        if 'RMOD' in Additional_Info:
            return 'RRU'  # Give priority to RMOD
        elif 'SMOD' in Additional_Info:
            return 'Baseband'
    return np.nan

# Apply the function for Nokia OEM and assign to the RRU/Baseband column
filtered_data.loc[filtered_data['OEM'] == "Nokia", 'RRU/Baseband'] = filtered_data['Additional_Info'].apply(classify_rru_baseband)

# Reapply classification where RRU/Baseband is still NaN
filtered_data.loc[(filtered_data['OEM'] == "Nokia") & (filtered_data['RRU/Baseband'].isna()),'RRU/Baseband'] = filtered_data['Additional_Info'].apply(classify_rru_baseband)



filtered_data['RRU/Baseband'] = filtered_data['RRU/Baseband'].fillna("#N/A")
#================================================================================working on RRU_Baseband_Fan_Temperature_Alarm=================================================================================================

filtered_data['RRU_Baseband_Fan_Temperature_Alarm'] = filtered_data.apply(
    lambda row: 'RRU Fan Alarm' if row['FAN/Temperature alarm'] == 'FAN alarm' and row['RRU/Baseband'] == 'RRU' 
    else 'Baseband Fan Alarm' if row['FAN/Temperature alarm'] == 'FAN alarm' and row['RRU/Baseband'] == 'Baseband'
    else 'RRU Temperature Alarm' if row['FAN/Temperature alarm'] == 'Temperature Alarm' and row['RRU/Baseband'] == 'RRU'
    else 'Baseband Temperature Alarm' if row['FAN/Temperature alarm'] == 'Temperature Alarm' and row['RRU/Baseband'] == 'Baseband'
    else 'Not Known',
    axis=1
)

#=================================================================================================working on Card Type=========================================================================================================

# Function to extract 'unitName' values
def extract_unit_name(text):
    match = re.search(r'unitName=([\w-]+)', str(text))
    return match.group(1) if match else None

# # Apply the function to the 'Additional_Info' column
# df['unitName'] = df['Additional_Info'].apply(extract_unit_name)

filtered_data.loc[filtered_data['OEM'] == "Nokia", 'Card Type'] = filtered_data['Additional_Info'].apply(extract_unit_name)

# Function to extract 'unitName' values
def extract_unit_name(text):
    match = re.search(r'FieldReplaceableUnit=([^-\s]+)', str(text))
    return match.group(1) if match else None

# # Apply the function to the 'Additional_Info' column
# df['unitName'] = df['Additional_Info'].apply(extract_unit_name)

filtered_data.loc[filtered_data['OEM'] == "Ericsson", 'Card Type'] = filtered_data['Additional_Info'].apply(extract_unit_name)

def extract_unit_name(text):
    match = re.search(r'Board Type=([\w-]+)', str(text))
    return match.group(1) if match else None

# # Apply the function to the 'Additional_Info' column
# df['unitName'] = df['Additional_Info'].apply(extract_unit_name)

filtered_data.loc[filtered_data['OEM'] == "Huawei", 'Card Type'] = filtered_data['Additional_Info'].apply(extract_unit_name)


def extract_unit_name(text):
    match = re.search(r'BOARDTYPE:([\w-]+)', str(text))
    return match.group(1) if match else None

# # Apply the function to the 'Additional_Info' column
# df['unitName'] = df['Additional_Info'].apply(extract_unit_name)

filtered_data.loc[filtered_data['OEM'] == "ZTE", 'Card Type'] = filtered_data['Additional_Info'].apply(extract_unit_name)



filtered_data['Card Type'] = filtered_data['Card Type'].fillna("#N/A")

#=================================================================================================working on Serial Number=========================================================================================================


def extract_serialno(text):
    match = re.search(r'serial_no=([\w-]+)', str(text))
    return match.group(1) if match else None

# # Apply the function to the 'Additional_Info' column
# df['unitName'] = df['Additional_Info'].apply(extract_unit_name)

filtered_data.loc[filtered_data['OEM'] == "Nokia", 'Serial No.'] = filtered_data['Additional_Info'].apply(extract_serialno)

def extract_serialno(text):
    match = re.search(r'Board serial number: ([\w-]+)', str(text))
    return match.group(1) if match else None

# # Apply the function to the 'Additional_Info' column
# df['unitName'] = df['Additional_Info'].apply(extract_unit_name)

filtered_data.loc[filtered_data['OEM'] == "ZTE", 'Serial No.'] = filtered_data['Additional_Info'].apply(extract_serialno)

#=================================================================================================working on MRBTS ID=========================================================================================================

filtered_data['MRBTS_ID'] = filtered_data['Alert_Key'].str.extract(r'(MRBTS-\d+)')
filtered_data['SiteID_MRBTS_ID'] = filtered_data['SITEID'] + "_" + filtered_data['MRBTS_ID']
filtered_data = pd.merge(filtered_data, df_amob_acoc, on='SiteID_MRBTS_ID', how='left')
filtered_data['Cabinet Type'] = filtered_data['Cabinet Type'].fillna("#N/A")
filtered_data['Date-Month'] = filtered_data['EMS Start Date'].dt.strftime('%d-%m')

#===============================================================================================working on Alarm Duration=========================================================================================================

# Calculate duration only if Status is "Resolved"
filtered_data["Alarm Duration"] = filtered_data.apply(
    lambda row: round((row["EMS End Time"] - row["EMS Start Time"]).total_seconds() / 60, 1) 
    if row["Status"] == "Resolved" else None, 
    axis=1
)

#===============================================================================================working on Alarm Type=========================================================================================================

# Define function to determine Alarm Type
def calculate_alarm_type(group):
    alarms = set(group["FAN/Temperature alarm"].unique())
    if {"Temperature Alarm", "FAN alarm"}.issubset(alarms):
        return "Fan+Temperature Alarm"
    else:
        return next(iter(alarms))  # Pop without removing from set

# Group by SITEID and Week Name, then apply function
alarm_types = filtered_data.groupby(["SITEID", "EMS Start Date"]).apply(calculate_alarm_type).reset_index(name="Alarm Type")

# Merge back with the original DataFrame
filtered_data = filtered_data.merge(alarm_types, on=["SITEID", "EMS Start Date"], how="left")

#======================================================================================================Calculating Consecutive_Days======================================================================================

#Get the latest date in the dataset
latest_date = filtered_data['EMS Start Date'].max()
second_max_date = filtered_data.loc[filtered_data['EMS Start Date'] < latest_date, 'EMS Start Date'].max()

#=============================================================================================Calculate calculate_alarm_period===============================================================================================


def calculate_alarm_period(group):
    group = group.sort_values('EMS Start Date', ascending=True)  # Sort ascending order
    
    # Initialize Consecutive_Days column
    group['Consecutive_Days'] = 1  
    
    # Iterate through rows to manually calculate consecutive days
    for i in range(1, len(group)):
        if group.iloc[i]['EMS Start Date'] == group.iloc[i - 1]['EMS Start Date']:
            # If EMS Start Date is same as previous, keep the same Consecutive_Days
            group.loc[group.index[i], 'Consecutive_Days'] = group.loc[group.index[i - 1], 'Consecutive_Days']
        elif (group.iloc[i]['EMS Start Date'] - group.iloc[i - 1]['EMS Start Date']).days == 1:
            # If the date is exactly +1 day, increase Consecutive_Days
            group.loc[group.index[i], 'Consecutive_Days'] = group.loc[group.index[i - 1], 'Consecutive_Days'] + 1
        else:
            # If there is a gap (missing date), reset to 1
            group.loc[group.index[i], 'Consecutive_Days'] = 1

    return group




# Apply function to each SITEID group
filtered_data = filtered_data.groupby(['Circle','SITEID','OEM','Card Type','Cabinet Type','Alarm Type'], group_keys=False).apply(calculate_alarm_period)



#=============================================================================================working on Alarm Occurrence Period================================================================================================
# Assign 'Alarm Occurrence Period' where Consecutive Days are 3 or more
filtered_data['Alarm Occurrence Period'] = np.where(filtered_data['Consecutive_Days'] >= 3, '≥ 3 Days', '#N/A')

# Final sorting to maintain order
filtered_data = filtered_data.sort_values(['SITEID', 'EMS Start Date'], ascending=[True, True])




#Filter data for the latest 3 days
latest1daysdump = filtered_data[filtered_data['EMS Start Date'] > (latest_date - pd.Timedelta(days=1))]
latest2daysdump = filtered_data[filtered_data['EMS Start Date'] >= (latest_date - pd.Timedelta(days=1))]
latest3daysdump = filtered_data[filtered_data['EMS Start Date'] >= (latest_date - pd.Timedelta(days=2))]
latest7daysdump = filtered_data[filtered_data['EMS Start Date'] >= (latest_date - pd.Timedelta(days=6))]


#============================================================================================Working on Sitewise_Backup_Daily==================================================================================
# Create Pivot Table
Sitewise_backup = filtered_data.pivot_table(
    index=['Circle', 'SITEID','OEM', 'Card Type','RRU/Baseband','Cabinet Type','Alarm Type'], 
    columns='EMS Start Date', 
    values='Date-Month',  # Use any column that exists for counting
    aggfunc= 'count',
    fill_value=np.nan,  # Fill missing values with 0
    margins=True,  # Add totals for both rows
    margins_name='Total'  # Name for the row and column totals
).reset_index()

# Format the column headers to show date in the desired format
Sitewise_backup.columns = [col.strftime('%d %b') if isinstance(col, pd.Timestamp) else col for col in Sitewise_backup.columns]


# Count unique EMS Start Dates per Circle and SITEID
site_date_counts = latest7daysdump.groupby(['Circle', 'SITEID','OEM','Card Type','Cabinet Type','Alarm Type'])['EMS Start Date'].nunique().reset_index(name='Count in Days/7 Days')
Sitewise_backup = Sitewise_backup.merge(site_date_counts, on=['Circle', 'SITEID','OEM','Card Type','Cabinet Type','Alarm Type'], how="left")

# Assuming Sitewise_backup and filtered_sites are already defined DataFrames
Sitewise_backup["Priority"] = Sitewise_backup["Count in Days/7 Days"].apply(
    lambda x: "P1Site(>=3/7days Alarm)" if x >= 3 else "Normal"
)

#=============================================>= 20 Alarms in two consecutive days and >= 3 Days Cosecutive=============================================

ems_dates = list(Sitewise_backup.columns[21:])
Sitewise_backup['>=20 Alarms in latest Two Days'] = np.where(
    (Sitewise_backup[ems_dates[0]] >= 20) & (Sitewise_backup[ems_dates[1]] >= 20), 
    "P1Site(>=20 Alarm)", "Normal"
)

ems_dates = list(Sitewise_backup.columns[20:])
Sitewise_backup['>= 3 Days Cosecutive'] = np.where(
    (Sitewise_backup[ems_dates[0]] > 0) & (Sitewise_backup[ems_dates[1]] > 0) & (Sitewise_backup[ems_dates[2]] > 0), 
    ">= 3 Days", "Normal"
)


#=======================================================================Alarm_Type=========================================================================

# Filter where the unique EMS Start Date count is at least 3
filtered_sites = Sitewise_backup[Sitewise_backup['Priority'] == 'P1Site(>=3/7days Alarm)']
# Count how many such SITEIDs exist per Circle
circle_counts = filtered_sites.groupby(['Circle'])['SITEID'].nunique().reset_index(name='P1Site(>=3/7days Alarm)')


greatertha20alarm = Sitewise_backup[Sitewise_backup['>=20 Alarms in latest Two Days'] == "P1Site(>=20 Alarm)"]
greatertha20alarm_grouped = greatertha20alarm.groupby(['Circle']).size().reset_index(name='>= 20 Alarms')
greatertha20alarm_grouped = greatertha20alarm_grouped.sort_values(by='>= 20 Alarms', ascending=False)

# Create Pivot Table
alarm_type = latest3daysdump.pivot_table(
    index=['Circle'], 
    columns='Alarm Type',
    values='SITEID',  # Use any column that exists for counting
    aggfunc=lambda x: x.nunique(),  # Aggregate function (count occurrences)
    fill_value=0,  # Fill missing values with 0
    # margins=True,  # Add totals for both rows
    # margins_name='Total'  # Name for the row and column totals
).reset_index()

# alarm_type = alarm_type.sort_values(by='Total', ascending=False)
alarm_type = pd.merge(alarm_type, circle_counts, on='Circle', how='left')
alarm_type = pd.merge(alarm_type, greatertha20alarm_grouped, on='Circle', how='left')
#alarm_type = pd.merge(alarm_type, weekly_count_grouped, on='Circle', how='left')

alarm_type['>= 20 Alarms'] = alarm_type['>= 20 Alarms'].fillna(0)

#=================================================================================High Temperature Table================================================================================

filtered_data_temp_alarm = latest7daysdump[latest7daysdump['FAN/Temperature alarm'] == 'Temperature Alarm']

# Pivot table
summary_temp_total = filtered_data_temp_alarm.pivot_table(
    index=['Circle', 'OEM'], 
    columns='EMS Start Date', 
    values='SITEID',  
    aggfunc='count',
    fill_value=0,  
    margins=True,  
    margins_name='Total'  # Adds "Total" row and column
).reset_index()

# Format the column headers to show dates in 'DD MMM' format
summary_temp_total.columns = [
    col.strftime('%d %b') if isinstance(col, pd.Timestamp) else col 
    for col in summary_temp_total.columns
]

# Separate 'Total' row
total_row = summary_temp_total[summary_temp_total['Circle'] == 'Total']
summary_temp_total = summary_temp_total[summary_temp_total['Circle'] != 'Total']

# Sort by 'Circle', 'OEM', and 'Total' column in descending order
summary_temp_total = summary_temp_total.sort_values(
    by=['Total'], 
    ascending=[False]
)

# Append 'Total' row at the bottom
summary_temp_total = pd.concat([summary_temp_total, total_row], ignore_index=True)
summary_temp_total = summary_temp_total.drop(columns=['Total'])






 # Create pivot table for Unique Sites count
summary_temp_unique = pd.pivot_table(
        filtered_data_temp_alarm,  # DataFrame containing the data
        values='SITEID',  # Column to count
        index=['Circle','OEM'],  # Row index (Circle)+
        columns='EMS Start Date',  # Columns (EMS Start Date)
        aggfunc=lambda x: x.nunique(),  # Aggregate function (count occurrences)
        fill_value = 0,  # Fill missing values with 0       
        margins=True,  # Add totals for both rows and columns
        margins_name='Total'  # Name for the row and column totals
        
    ).reset_index()

# Format the column headers to show date in the desired format
summary_temp_unique.columns = [col.strftime('%d %b') if isinstance(col, pd.Timestamp) else col for col in summary_temp_unique.columns]

# Separate 'Total' row
total_row = summary_temp_unique[summary_temp_unique['Circle'] == 'Total']
summary_temp_unique = summary_temp_unique[summary_temp_unique['Circle'] != 'Total']

# Sort by 'Circle', 'OEM', and 'Total' column in descending order
summary_temp_unique = summary_temp_unique.sort_values(
    by=['Total'], 
    ascending=[False]
)

# Append 'Total' row at the bottom
summary_temp_unique = pd.concat([summary_temp_unique, total_row], ignore_index=True)
summary_temp_unique = summary_temp_unique.drop(columns=['Total'])
#================================================================================Fan Alarm Table===========================================================================================
filtered_data_fan_alarm = latest7daysdump[latest7daysdump['FAN/Temperature alarm'] == 'FAN alarm']

summary_fan_total = filtered_data_fan_alarm.pivot_table(
    index=['Circle', 'OEM'], 
    columns='EMS Start Date', 
    values='SITEID',  # Use any column that exists for counting
    aggfunc= 'count',
    fill_value = 0,  # Fill missing values with 0
    margins=True,  # Add totals for both rows
    margins_name='Total'  # Name for the row and column totals
).reset_index()

# Format the column headers to show date in the desired format
summary_fan_total.columns = [col.strftime('%d %b') if isinstance(col, pd.Timestamp) else col for col in summary_fan_total.columns]


# Separate 'Total' row
total_row = summary_fan_total[summary_fan_total['Circle'] == 'Total']
summary_fan_total = summary_fan_total[summary_fan_total['Circle'] != 'Total']

# Sort by 'Circle', 'OEM', and 'Total' column in descending order
summary_fan_total = summary_fan_total.sort_values(
    by=['Total'], 
    ascending=[False]
)

# Append 'Total' row at the bottom
summary_fan_total = pd.concat([summary_fan_total, total_row], ignore_index=True)
summary_fan_total = summary_fan_total.drop(columns=['Total'])


 # Create pivot table for Unique Sites count
summary_fan_unique = pd.pivot_table(
        filtered_data_fan_alarm,  # DataFrame containing the data
        values='SITEID',  # Column to count
        index=['Circle','OEM'],  # Row index (Circle)+
        columns='EMS Start Date',  # Columns (EMS Start Date)
        aggfunc=lambda x: x.nunique(),  # Aggregate function (count occurrences)
        fill_value = 0,  # Fill missing values with 0       
        margins=True,  # Add totals for both rows and columns
        margins_name='Total'  # Name for the row and column totals
        
    ).reset_index()

# Format the column headers to show date in the desired format
summary_fan_unique.columns = [col.strftime('%d %b') if isinstance(col, pd.Timestamp) else col for col in summary_fan_unique.columns]


# Separate 'Total' row
total_row = summary_fan_unique[summary_fan_unique['Circle'] == 'Total']
summary_fan_unique = summary_fan_unique[summary_fan_unique['Circle'] != 'Total']

# Sort by 'Circle', 'OEM', and 'Total' column in descending order
summary_fan_unique = summary_fan_unique.sort_values(
    by=['Total'], 
    ascending=[False]
)

# Append 'Total' row at the bottom
summary_fan_unique = pd.concat([summary_fan_unique, total_row], ignore_index=True)
summary_fan_unique = summary_fan_unique.drop(columns=['Total'])
#==================================================================================Consecutive 3 Days Table==============================================================================
grth3dayscons = latest7daysdump[latest7daysdump['Alarm Occurrence Period'] == '≥ 3 Days']
 # Create pivot table for Unique Sites count
summary_grth3daycons = pd.pivot_table(
        grth3dayscons,  # DataFrame containing the data
        values='SITEID',  # Column to count
        index=['Circle','OEM'],  # Row index (Circle)+
        columns='EMS Start Date',  # Columns (EMS Start Date)
        aggfunc=lambda x: x.nunique(),  # Aggregate function (count occurrences)
        fill_value = 0,  # Fill missing values with 0       
        margins=True,  # Add totals for both rows and columns
        margins_name='Total'  # Name for the row and column totals
        
    ).reset_index()

# Format the column headers to show date in the desired format
summary_grth3daycons.columns = [col.strftime('%d %b') if isinstance(col, pd.Timestamp) else col for col in summary_grth3daycons.columns]

# Separate 'Total' row
total_row = summary_grth3daycons[summary_grth3daycons['Circle'] == 'Total']
summary_grth3daycons = summary_grth3daycons[summary_grth3daycons['Circle'] != 'Total']

# Sort by 'Circle', 'OEM', and 'Total' column in descending order
summary_grth3daycons = summary_grth3daycons.sort_values(
    by=['Total'], 
    ascending=[False]
)

# Append 'Total' row at the bottom
summary_grth3daycons = pd.concat([summary_grth3daycons, total_row], ignore_index=True)
summary_grth3daycons = summary_grth3daycons.drop(columns=['Total'])
#============================defining for mail body=======================================

# Function to get top 5 circles for each category
def get_top_circles(df, column):
    return ", ".join(df.nlargest(5, column)["Circle"])

# Finding top 5 circles for each category
top_fan_alarm = get_top_circles(alarm_type, "FAN alarm")
top_temp_alarm = get_top_circles(alarm_type, "Temperature Alarm")
top_fan_temp_alarm = get_top_circles(alarm_type, "Fan+Temperature Alarm")
top_p1site_alarm = get_top_circles(alarm_type, "P1Site(>=3/7days Alarm)")
top_20_alarm = get_top_circles(alarm_type, ">= 20 Alarms")



#=================================================================Dashboard High Temp=======================================================================================
filtered_data = filtered_data.drop(columns=['Date-Month'])
# Save the data to Excel

# Define output file path
#==========================================================================================

#========================================Printing required Dataframes==============================================
report_date = combined_data_high_temp['EMS Start Date'].max().strftime('%d-%m-%Y')
output_folder = r"D:\Automation\High Temp Analysis\output"
#yesterday = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime("%Y%m%d")
output_file = os.path.join(output_folder, f'RAN_High_Temperature_and_FAN_Alarms_{report_date}.xlsx')


# Create output folder if it doesn't exist
os.makedirs(output_folder, exist_ok=True)


# Write to Excel
with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:     
    
    alarm_type.to_excel(writer, sheet_name="Summary", index=False, startrow=2, startcol=1)
    summary_grth3daycons.to_excel(writer, sheet_name="Summary", index=False, startrow=2, startcol=8)
    summary_temp_total.to_excel(writer, sheet_name="Summary", index=False, startrow=37, startcol=1)
    summary_temp_unique.to_excel(writer, sheet_name="Summary", index=False, startrow=37, startcol=11)
    summary_fan_total.to_excel(writer, sheet_name="Summary", index=False, startrow=73, startcol=1)
    summary_fan_unique.to_excel(writer, sheet_name="Summary", index=False, startrow=73, startcol=11)
    Sitewise_backup.to_excel(writer, sheet_name="Sitewise_Backup_Daily", index=False)
    latest3daysdump.to_excel(writer, sheet_name="Dump_Latest_3_Days", index=False) 
    filtered_data.to_excel(writer, sheet_name="Filtered_Alarm_Data", index=False)
   

    # Apply formatting
    workbook  = writer.book
    worksheet = writer.sheets["Summary"]

    merge_format = workbook.add_format({
        'bold': True,
        'align': 'center',
        'valign': 'vcenter',
        'border': 1
    })

    # Merge cells with labels
    worksheet.merge_range(1, 1, 1, 6, 'Alarm Type Category', merge_format)
    worksheet.merge_range(1, 8, 1, 16, '>= 3 Consecutive Days Site Count', merge_format)
    worksheet.merge_range(36, 1, 36, 9, 'Temperature Alarm Count', merge_format)
    worksheet.merge_range(36, 11, 36, 19, 'Temperature Daily Unique Sites', merge_format)
    worksheet.merge_range(72, 1, 72, 9, 'FAN Alarm Count', merge_format)
    worksheet.merge_range(72, 11, 72, 19, 'FAN Alarm Daily Unique Sites', merge_format)


print("Data saved successfully to multiple sheets in:",output_file)

# #=================================sending Mail=====================================

# # Read recipients from Excel
# file_path_receipient = r"D:\Automation\High Temp Analysis\Receipient.xlsx"
# try:
#     receipient_df = pd.read_excel(file_path_receipient)
#     to_recipients = receipient_df["to"].dropna().tolist()  # List of 'To' recipients
#     cc_recipients = receipient_df["cc"].dropna().tolist()  # List of 'CC' recipients
# except Exception as e:
#     print(f"❌ Error reading Excel file: {e}")
#     sys.exit()

# # Define sender email
# from_email = "IN_D_Engineering_NIfi_Reports@airtel.com"

# # Initialize Outlook
# try:
#     outlook = win32com.client.Dispatch("Outlook.Application")
#     namespace = outlook.GetNamespace("MAPI")
# except Exception as e:
#     print(f"❌ Error initializing Outlook: {e}")
#     sys.exit()

# # Get today's date for email subject
# today_date = datetime.now().strftime("%d%m%Y")
# yesterday_date = (datetime.now() - timedelta(days=1)).strftime("%d%m%Y")


# email_subject = f"FAN and Temperature Alarms - {yesterday_date}"
# email_body = f"""
# Dear Team,

# Please find Key observations from the latest alarm data:
# 🔹 Top 5 Circles in FAN Alarm: {top_fan_alarm}
# 🔹 Top 5 Circles in Temperature Alarm: {top_temp_alarm}
# 🔹 Top 5 Circles in Fan+Temperature Alarms: {top_fan_temp_alarm} 
# 🔹 Top 5 Circles in P1Site (≥3/7 days Alarm): {top_p1site_alarm}
# 🔹 Top 5 Circles in ≥20 Alarms: {top_20_alarm}

# Regards,  
# Central Team FSO
# """

# #output_file_path = r"D:\Automation\High Temp Analysis\output\RAN_High_Temperature_and_FAN_Alarms.xlsx"  # Adjust path

# try:
#     # Create a new email
#     mail = outlook.CreateItem(0)
#     mail.To = ";".join(to_recipients) if to_recipients else ""  # Join multiple recipients with ";"
#     mail.CC = ";".join(cc_recipients) if cc_recipients else ""  # Join multiple CC recipients
#     mail.Subject = email_subject
#     mail.Body = email_body

#     # Specify the sender email
#     mail.SentOnBehalfOfName = from_email  # Option 1 (if account has permission)

#     # Alternative approach (set the sender account explicitly)
#     for account in outlook.Session.Accounts:
#         if account.SmtpAddress == from_email:
#             mail._oleobj_.Invoke(*(64209, 0, 8, 0, account))  # Option 2 (SendUsingAccount)

#     # Attach the file
#     mail.Attachments.Add(output_file)
#     mail.Send()  # Send the email

#     print(f"📧 Email sent successfully from {from_email}")

# except Exception as e:
#     print(f"❌ Error sending email: {e}")

# # Release Outlook objects
# del outlook, namespace

# print("✅ Email process completed successfully!")
# #=================================================================================================End of Code=========================================================================================================



# # #==========================================================================================
# # output_file_path = r"D:\Automation\High Temp Analysis\output\RAN_High_Temperature_and_FAN_Alarms.xlsx"

# # # Try loading existing workbook to preserve formatting
# # try:
# #     book = load_workbook(output_file_path)
# #     # Delete "Sitewise_Backup_Daily" sheet if it exists (for overwriting)
# #     sheets_to_remove = ["Sitewise_Backup_Daily", "Dump_Latest_3_Days"]
# #     for sheet_name in sheets_to_remove:
# #         if sheet_name in book.sheetnames:
# #             std = book[sheet_name]
# #             book.remove(std)

# #     book.save(output_file_path)  # Save after removal

# #     writer = pd.ExcelWriter(output_file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay")
# #     writer._book = book  # Correct way to assign the workbook
# #     #writer._sheets = {sheet.title: sheet for sheet in book.worksheets}
# # except FileNotFoundError:
# #     writer = pd.ExcelWriter(output_file_path, engine="openpyxl", mode="w")

# # # Write DataFrames to existing file without changing formatting
# # alarm_type.to_excel(writer, sheet_name="Summary", index=False, startrow=2, startcol=1)
# # summary_grth3daycons.to_excel(writer, sheet_name="Summary", index=False, startrow=2, startcol=8)
# # summary_temp_total.to_excel(writer, sheet_name="Summary", index=False, startrow=37, startcol=1)
# # summary_temp_unique.to_excel(writer, sheet_name="Summary", index=False, startrow=37, startcol=11)
# # summary_fan_total.to_excel(writer, sheet_name="Summary", index=False, startrow=73, startcol=1)
# # summary_fan_unique.to_excel(writer, sheet_name="Summary", index=False, startrow=73, startcol=11)
# # Sitewise_backup.to_excel(writer, sheet_name="Sitewise_Backup_Daily", index=False)
# # latest3daysdump.to_excel(writer, sheet_name="Dump_Latest_3_Days", index=False) 

# # writer.close()

# # # Apply formatting using openpyxl
# # book = load_workbook(output_file_path)
# # ws = book["Summary"]


# # # Define thick border
# # thin_border = Border(
# #     left=Side(style='thin'),
# #     right=Side(style='thin'),
# #     top=Side(style='thin'),
# #     bottom=Side(style='thin')
# # )

# # # Define red background and white bold text
# # fill = PatternFill(start_color="FF0000", fill_type="solid")  # Red background
# # font = Font(color="FFFFFF", bold=True)  # White bold text

# # for col in list(range(2, 8)) + list(range(9, 18)):  # Excludes column H (8)
# #     cell = ws.cell(row=3, column=col)
# #     cell.fill = fill
# #     cell.font = font
# #     cell.border = thin_border


# # for col in  list(range(2, 11)) + list(range(12, 21)):
# #     cell = ws.cell(row=38, column=col)
# #     cell.fill = fill
# #     cell.font = font
# #     cell.border = thin_border

# # for col in  list(range(2, 11)) + list(range(12, 21)):  
# #     cell = ws.cell(row=74, column=col)
# #     cell.fill = fill
# #     cell.font = font
# #     cell.border = thin_border

# # book.save(output_file_path)  # Save the final formatting


