import openpyxl
import os

# ===== CONFIGURATION =====
# File Path 
folder_path = r"C:\Users\b0335496\Downloads\RAN_High_Temperature_and_FAN_Alarms_01-12-2025"
sheets_to_delete = ["Sitewise_Backup_Daily", "Summary"]

try:
    # 1. Folder check karein
    if not os.path.exists(folder_path):
        print(f"❌ Error: Path nahi mila -> {folder_path}")
    else:
        # 2. Create the Ecel File List
        files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
        
        if not files:
            print("📂 Is folder mein koi .xlsx file nahi mili.")
        else:
            print(f"🔄 Total {len(files)} files mili. Clean-up shuru ho raha hai...\n")

        for file_name in files:
            file_path = os.path.join(folder_path, file_name)
            
            try:
                # 3. Workbook Load
                # data_only=False taaki formulas distrub na hon
                workbook = openpyxl.load_workbook(file_path)
                file_modified = False

                # 4. Sheets Check and Delete
                for sheet_name in sheets_to_delete:
                    if sheet_name in workbook.sheetnames:
                        # Workbook mein kam se kam 1 sheet honi zaroori hai
                        if len(workbook.sheetnames) > 1:
                            std = workbook[sheet_name]
                            workbook.remove(std)
                            print(f"✅ {file_name} -> Sheet '{sheet_name}' deleted.")
                            file_modified = True
                        else:
                            print(f"⚠️ {file_name} -> '{sheet_name}' delete nahi ki kyunki ye last sheet hai.")
                
                # 5. Save if changes were made
                if file_modified:
                    workbook.save(file_path)
                    workbook.close()
                else:
                    print(f"ℹ️ {file_name} -> Sheets nahi mili, skip kiya gaya.")

            except Exception as file_error:
                print(f"❌ Error in file {file_name}: {file_error}")

    print("\n✨ Clean-up process complete!")

except Exception as e:
    print(f"❌ Main process error: {str(e)}")