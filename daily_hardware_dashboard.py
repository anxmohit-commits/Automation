from itertools import count
from click import group
import pandas as pd
import numpy as np
import glob
import os
import re 
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side ,Alignment
import win32com.client
import win32timezone
import os
from datetime import datetime, timedelta
from pathlib import Path


#================================================================================
# ✅ Fetching Last 7 Days' Reports from Outlook
#================================================================================

# Define the folder where attachments will be saved
#save_path = r"D:\Daily_Hardware_Alarm_Dump"

save_paths = {
    "RAN Daily Hardware Alarms": Path(r"D:\Daily_Hardware_Alarm_Dump"),
    "Microwave Hardware daily report": Path(r"D:\Microwave Hardware daily report")
}

# Ensure save directories exist
for path in save_paths.values():
    os.makedirs(path, exist_ok=True)

# Connect to Outlook
try:
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
except Exception as e:
    print(f"❌ Error connecting to Outlook: {e}")
    sys.exit()

# Select the Inbox → Required to Work folder
try:
    inbox = outlook.GetDefaultFolder(6)  # Inbox
    target_folder = inbox.Folders("Required to Work")
except Exception as e:
    print("❌ Error: 'Required to Work' folder not found!")
    sys.exit()

# Filter and sort messages
messages = target_folder.Items
messages.Sort("[ReceivedTime]", True)
three_days_ago = (datetime.now() - timedelta(days=3)).strftime('%m/%d/%Y')
messages = messages.Restrict(f"[ReceivedTime] >= '{three_days_ago}'")

# Check if any emails exist
if messages.Count == 0:
    print("❌ No emails received in the last 3 days. Exiting!")
    sys.exit()

found = False
for message in messages:
    try:
        subject = message.Subject
        received_time = message.ReceivedTime
        print(f"📧 Checking email: {subject} | Date: {received_time.strftime('%Y-%m-%d')}")

        for keyword, folder_path in save_paths.items():
            if keyword in subject:
                print(f"✅ Found matching email for: {keyword}")
                if message.Attachments.Count > 0:
                    for attachment in message.Attachments:
                        if attachment.FileName.endswith(".xlsx"):
                            mail_date = received_time.strftime('%Y%m%d')
                            base_name, ext = os.path.splitext(attachment.FileName)
                            new_file_name = f"{base_name}_{mail_date}{ext}"
                            file_path = os.path.join(folder_path, new_file_name)

                            try:
                                attachment.SaveAsFile(file_path)
                                print(f"📂 Attachment saved: {file_path}")
                                found = True
                            except Exception as e:
                                print(f"❌ Error saving attachment: {e}")
                else:
                    print("⚠️ Email has no attachments.")
                break  # No need to check other keywords once matched

    except Exception as e:
        print(f"⚠️ Error processing email: {e}")

if not found:
    print("❌ No matching emails with attachments found in the last 3 days.")

# Cleanup
del messages, target_folder, inbox, outlook

print("🚀 Proceeding to the next script execution...")


# #=============================================================================Reading hardware Alarm Files============================================================
# Folder path containing the files
folder_path = "D:/Daily_Hardware_Alarm_Dump"

# # Get all .xlsx files in the folder
# file_paths = glob.glob(os.path.join(folder_path, "*.xlsx"))
# file_paths.sort(key=os.path.getmtime, reverse=True)



# Get all .xlsx files and sort by filename (reverse for latest files first)
file_paths = glob.glob(os.path.join(folder_path, "*.xlsx"))
file_paths.sort(reverse=True)  # Alphabetically descending (latest file names first)



# Get only the latest 9 files
latest_files = file_paths[:17]

# Columns to load from each file
columns_to_load = ['Row Number', 'Circle', 'SITEID', 'Alarm Name', 'Node', 'Technology', 
                   'EMS Start Time', 'EMS End Time', 'Status', 'Alert Key', 'Additional Info', 'Agent']

# Initialize a list to hold data from each file
all_data = []

# Read each file and append the data to the list
for file_path in latest_files:
    try:
        file_size = os.path.getsize(file_path) / (1024 * 1024)  # Convert bytes to MB
        print(f"Reading file: {os.path.basename(file_path)} | Size: {file_size:.2f} MB")
        
        data = pd.read_excel(file_path, usecols=columns_to_load)
        all_data.append(data)
    except Exception as e:
        print(f"Error reading {file_path}: {e}")

# Concatenate all data into a single DataFrame
if all_data:
    combined_data_high_temp = pd.concat(all_data, ignore_index=True)
    print("Data successfully combined.")
else:
    combined_data_high_temp = pd.DataFrame(columns=columns_to_load)
    print("No valid data found.")

#===========================================================================processing High Temp Fan========================================================================
combined_data_high_temp = combined_data_high_temp.dropna(subset=['EMS Start Time'])
combined_data_high_temp.rename(columns={'Alarm Name': 'Alarm_Name','Additional Info': 'Additional_Info','Alert Key': 'Alert_Key'}, inplace=True)
#combined_data_high_temp = combined_data_high_temp.dropna(subset=['EMS Start Time'])
combined_data_high_temp['EMS Start Date'] = combined_data_high_temp['EMS Start Time'].dt.date
combined_data_high_temp['EMS Start Date'] = pd.to_datetime(combined_data_high_temp['EMS Start Date'], errors='coerce')
# Get today's datee
today_date = pd.to_datetime('today').normalize()
# Filter out rows where 'EMS Start Date' is today
combined_data_high_temp = combined_data_high_temp[combined_data_high_temp['EMS Start Date'] < today_date]
combined_data_high_temp = combined_data_high_temp[
    ~combined_data_high_temp['SITEID'].str.contains("Unknown", case=False, na=False)
]
combined_data_high_temp['Week Number'] = combined_data_high_temp['EMS Start Date'].dt.isocalendar().week
combined_data_high_temp['Week Name'] = 'WK-' + combined_data_high_temp['EMS Start Date'].dt.isocalendar().week.astype(str)

# Extract unique dates and drop NaT values
unique_dates = combined_data_high_temp['EMS Start Date'].unique()
# Sort the unique dates and format them
readable_dates = [date.strftime('%d-%m-%Y') for date in sorted(unique_dates)]
# Print the unique, readable dates
print("Unique EMS Start Dates:")
for date in readable_dates:
    print(date)

#start_date_input = "23-01-2025"

try:
    # Convert start date to datetime
    #start_date = pd.to_datetime(start_date_input, format='%d-%m-%Y', errors='coerce')

    # Ensure 'EMS Start Date' is in datetime format
    combined_data_high_temp['EMS Start Date'] = pd.to_datetime(combined_data_high_temp['EMS Start Date'], errors='coerce')

    # Get max EMS Start Date
    
    end_date = combined_data_high_temp['EMS Start Date'].max()
    start_date = end_date - pd.Timedelta(days=15)

    # Check if dates are valid
    if pd.isna(start_date) or pd.isna(end_date):
        print("Invalid date format or missing data.")
    elif start_date > end_date:
        print("Start date cannot be later than the max EMS Start Date.")
    else:
        # Filter data
        filtered_data = combined_data_high_temp[
            (combined_data_high_temp['EMS Start Date'] >= start_date) &
            (combined_data_high_temp['EMS Start Date'] <= end_date)
        ]

        print(f"Filtered data between {start_date.strftime('%d-%m-%Y')} and {end_date.strftime('%d-%m-%Y')}:")

except Exception as e:
    print(f"An error occurred: {e}")

#===========================================================Working on OEM Column====================================================================================


#Creating new column OEM
filtered_data = filtered_data.copy() 
filtered_data.loc[:,'OEM'] = filtered_data['Agent'].apply(
    lambda x: 'Ericsson' if 'Ericsson' in x else
              'Nokia' if 'Nokia' in x else
              'Huawei' if 'Huawei' in x else
              'ZTE' if 'ZTE' in x else
              'SAMSUNG' if 'SAMSUNG' in x else
              'IPA' if 'IPA' in x else
              'Unknown'
)
filtered_data = filtered_data.copy() 
filtered_data.rename(columns={'Alert Key': 'Alert_Key'}, inplace=True)


#==============================================Extracting Ericsson Card Type======================================================

def classify_card_type(Alert_Key):
    if isinstance(Alert_Key, str):
        if 'AAS' in Alert_Key:
            return 'AAS'
        elif 'Unit=4' in Alert_Key:
            return 'BBU'
        elif 'Clock' in Alert_Key:
            return 'GPS'
        elif 'RiLink' in Alert_Key:
            return 'CPRI/SFP'
        elif 'RRU' in Alert_Key:
            return 'RRU'
        else:
            return np.nan
    return np.nan

filtered_data = filtered_data.copy() 
filtered_data.loc[filtered_data['OEM'] == "Ericsson", 'Card Type'] = filtered_data['Alert_Key'].apply(classify_card_type)

#==============================================Extracting Samsung Card Type======================================================
def classify_card_type(Alert_Key):
    if isinstance(Alert_Key, str):
        if 'CPRI' in Alert_Key:
            return 'CPRI'
        elif 'RRH' in Alert_Key:
            return 'RRH'
        else:
            return np.nan
    return np.nan

filtered_data = filtered_data.copy() 
filtered_data.loc[filtered_data['OEM'] == "SAMSUNG", 'Card Type'] = filtered_data['Alert_Key'].apply(classify_card_type)


#==============================================Extracting ZTE Card Type======================================================

def classify_card_type(Additional_Info):
    if isinstance(Additional_Info, str):
        if 'RRU' in Additional_Info:
            return 'RRU'
        elif 'Clock' in Additional_Info:
            return 'GPS'
        elif 'Optical' in Additional_Info:
            return 'CPRI'
        else:
            return np.nan
    return np.nan

filtered_data = filtered_data.copy() 
filtered_data.loc[filtered_data['OEM'] == "ZTE", 'Card Type'] = filtered_data['Additional_Info'].apply(classify_card_type)


#==============================================Extracting Nokia Card Type======================================================

# Function to extract 'unitName' values
def extract_unit_name(text):
    match = re.search(r'unitName=([\w-]+)', str(text))
    return match.group(1) if match else None

# # Apply the function to the 'Additional_Info' column
# df['unitName'] = df['Additional_Info'].apply(extract_unit_name)

filtered_data.loc[filtered_data['OEM'] == "Nokia", 'Card Type'] = filtered_data['Additional_Info'].apply(extract_unit_name)

# filtered_data['Card Type'] = filtered_data['Alert_Key'].apply(classify_card_type)
filtered_data = filtered_data[filtered_data['Card Type'].notna() & (filtered_data['Card Type'] != '')]
filtered_data = filtered_data[~filtered_data['SITEID'].str.contains('TEST', case=False, na=False)]

#=================================================================Extracting Alarm Name for all OEM===================================

# Extract Alarm Name using split and taking the first part before `/` or `|`
#filtered_data["Extracted_Alarm_Name"] = filtered_data["Additional_Info"].str.split(r'[/|]').str[0]
filtered_data.loc[filtered_data["OEM"] == "Nokia", "Extracted_Alarm_Name"] = filtered_data["Additional_Info"].str.split(r'[/|;]').str[0]
filtered_data.loc[filtered_data["OEM"].isin(["Ericsson", "SAMSUNG", "ZTE"]), "Extracted_Alarm_Name"] = filtered_data["Alarm_Name"]
filtered_data['Date-Month'] = filtered_data['EMS Start Date'].dt.strftime('%d-%m')
filtered_data["Extracted_Alarm_Name"] = filtered_data["Extracted_Alarm_Name"].str.replace("5G ", "", regex=False)

#==================================================================Defining Category column============================================

# Define conditions
condition_cpri_sfp = filtered_data['Extracted_Alarm_Name'].isin([
    "LINK DEGRADED",
    "LINK FAILURE",
    "Increased BER detected on the optical connector",
    "Increased BER detected on the optical connection to Radio Module",
    "Receiving failure in Optical Interface",
    "IMP OPTIC-DIAGNOSIS-ABNORMAL",
    "RRH DIGITAL-INPUT-LOW"
])

condition_gps = filtered_data['Extracted_Alarm_Name'] == "TIMESYNCIO REFERENCE FAILED"

condition_hw_fault = filtered_data['Extracted_Alarm_Name'].isin([
    "LINEARIZATION DISTURBANCE PERFORMANCE DEGRADED",
    "HW PARTIAL FAULT",
    "FRU GENERAL PROBLEM",
    "HW FAULT",
    "TX out of order",
    "RF Module gain adjusting failure",
    "RF BB bus configuration error",
    "RRU Power Abnormal Alarm",
    "Internal fault"
])

# Define category values
categories = ["CPRI/SFP/Consumable", "GPS", "Hardware fault"]

# Apply conditions
filtered_data['Category'] = np.select(
    [condition_cpri_sfp, condition_gps, condition_hw_fault],
    categories,
    default=""
)
filtered_data['Category'] = filtered_data['Category'].fillna("#N/A")
#==========================================================calculating consecutive days=======================================

def calculate_alarm_period(group):
    group = group.sort_values('EMS Start Date', ascending=True)  # Sort ascending order
    
    # Initialize Consecutive_Days column
    group['Consecutive_Days'] = 1  
    
    # Iterate through rows to manually calculate consecutive days
    for i in range(1, len(group)):
        if group.iloc[i]['EMS Start Date'] == group.iloc[i - 1]['EMS Start Date']:
            # If EMS Start Date is same as previous, keep the same Consecutive_Days
            group.loc[group.index[i], 'Consecutive_Days'] = group.loc[group.index[i - 1], 'Consecutive_Days']
        elif (group.iloc[i]['EMS Start Date'] - group.iloc[i - 1]['EMS Start Date']).days == 1:
            # If the date is exactly +1 day, increase Consecutive_Days
            group.loc[group.index[i], 'Consecutive_Days'] = group.loc[group.index[i - 1], 'Consecutive_Days'] + 1
        else:
            # If there is a gap (missing date), reset to 1
            group.loc[group.index[i], 'Consecutive_Days'] = 1

    return group

filtered_data = filtered_data.groupby(
    ['Circle', 'OEM', 'SITEID', 'Extracted_Alarm_Name', 'Card Type', 'Category'],
    group_keys=False
).apply(calculate_alarm_period)

#===========================================================Defining Alarm Occurance Period=================================================

# Assign 'Alarm Occurrence Period' where Consecutive Days are 3 or more
filtered_data['Alarm Occurrence Period'] = np.where(filtered_data['Consecutive_Days'] >= 3, '≥ 3 Days', 'Normal')

# Final sorting to maintain order
filtered_data = filtered_data.sort_values(['SITEID', 'EMS Start Date'], ascending=[True, True])




latest_date = filtered_data['EMS Start Date'].max()

#Filter data for the latest 3 days
latest1daysdump = filtered_data[filtered_data['EMS Start Date'] > (latest_date - pd.Timedelta(days=1))]
latest2daysdump = filtered_data[filtered_data['EMS Start Date'] >= (latest_date - pd.Timedelta(days=1))]
latest3daysdump = filtered_data[filtered_data['EMS Start Date'] >= (latest_date - pd.Timedelta(days=2))]
latest7daysdump = filtered_data[filtered_data['EMS Start Date'] >= (latest_date - pd.Timedelta(days=6))]
grth3dayscons = latest7daysdump[latest7daysdump['Alarm Occurrence Period'] == '≥ 3 Days']
filtered_ericsson = filtered_data[filtered_data['OEM'] == 'Ericsson']
filtered_Nokia = filtered_data[filtered_data['OEM'] == 'Nokia']

#======================================================Working on Sitewise_Backup_Daily==================================================================================

# Count unique EMS Start Dates per Circle and SITEID
site_date_counts = latest7daysdump.groupby(['Circle', 'OEM', 'SITEID', 'Extracted_Alarm_Name', 'Card Type', 'Category'])['EMS Start Date'].nunique().reset_index(name='Count in Days/7 Days')

# Filter where the unique EMS Start Date count is at least 3
filtered_sites = site_date_counts[site_date_counts['Count in Days/7 Days'] >= 3]
# Count how many such SITEIDs exist per Circle
circle_counts = filtered_sites.groupby(['Circle'])['SITEID'].nunique().reset_index(name='P1Site(>=3/7days Alarm)')






# Create Pivot Table
Sitewise_backup = filtered_data.pivot_table(
    index=['Circle','OEM','SITEID','Extracted_Alarm_Name', 'Card Type','Category'], 
    columns='EMS Start Date', 
    values='Date-Month',  # Use any column that exists for counting
    aggfunc= 'count',
    fill_value=np.nan,  # Fill missing values with 0
    margins=True,  # Add totals for both rows
    margins_name='Total'  # Name for the row and column totals
).reset_index()

# Format the column headers to show date in the desired format
Sitewise_backup.columns = [col.strftime('%d %b') if isinstance(col, pd.Timestamp) else col for col in Sitewise_backup.columns]
Sitewise_backup = Sitewise_backup[Sitewise_backup['Circle'] != 'Total']

Sitewise_backup = Sitewise_backup.merge(site_date_counts, on=['Circle', 'OEM','SITEID','Extracted_Alarm_Name','Card Type','Category'], how="left")

# Assuming Sitewise_backup and filtered_sites are already defined DataFrames
Sitewise_backup["Priority"] = Sitewise_backup["Count in Days/7 Days"].apply(
    lambda x: "P1Site(>=3/7days Alarm)" if x >= 3 else "Normal"
)


# Define conditions and corresponding values
conditions = [
    Sitewise_backup["Count in Days/7 Days"].isin([1, 2]),  # If value is 1 or 2
    Sitewise_backup["Count in Days/7 Days"].isin([3, 4]),  # If value is 3 or 4
    Sitewise_backup["Count in Days/7 Days"] >= 5           # If value is 5 or more
]

values = ["1-2 days/7days", "3-4 days/7days", ">=5 days/7days"]

# Apply the classification
Sitewise_backup["Site Action Priority"] = np.select(conditions, values, default="Normal")




#=============================================>= 20 Alarms in two consecutive days and >= 3 Days Cosecutive=============================================

ems_dates = list(Sitewise_backup.columns[20:])
Sitewise_backup['>=20 Alarms in latest Two Days'] = np.where(
    (Sitewise_backup[ems_dates[0]] >= 20) & (Sitewise_backup[ems_dates[1]] >= 20), 
    "P1Site(>=20 Alarm)", "Normal"
)

ems_dates = list(Sitewise_backup.columns[19:])
Sitewise_backup['>= 3 Days Cosecutive'] = np.where(
    (Sitewise_backup[ems_dates[0]] > 0) & (Sitewise_backup[ems_dates[1]] > 0) & (Sitewise_backup[ems_dates[2]] > 0), 
    ">= 3 Days", "Normal"
)


#========================================================================================================================================================
# Count unique EMS Start Dates for each SITEID
site_priority = latest7daysdump.groupby("SITEID")["EMS Start Date"].nunique().reset_index()
site_priority.rename(columns={"EMS Start Date": "Unique EMS Start Count"}, inplace=True)

# Assign Site Action Priority based on unique count of EMS Start Dates
def assign_priority(count):
    if count <= 2:
        return "1-2 days/7days"
    elif 3 <= count <= 4:
        return "3-4 days/7days"
    elif count >= 5:
        return ">=5 days/7days"
    else:
        return "Unknown"
    
# Merge the unique count back into latest7daysdump
latest7daysdump = latest7daysdump.merge(site_priority, on="SITEID", how="left")
# Now apply the function correctly
latest7daysdump["Site Action Priority"] = latest7daysdump["Unique EMS Start Count"].apply(assign_priority)



#==================================================Hardware_Summary based on last 7 days data=============================================


summary_hw_total = latest7daysdump.pivot_table(
    index=['Circle', 'OEM'], 
    columns='EMS Start Date', 
    values='SITEID',  # Use any column that exists for counting
    aggfunc= 'count',
    fill_value = 0,  # Fill missing values with 0
    margins=True,  # Add totals for both rows
    margins_name='Total'  # Name for the row and column totals
).reset_index()

# Format the column headers to show date in the desired format
summary_hw_total.columns = [col.strftime('%d %b') if isinstance(col, pd.Timestamp) else col for col in summary_hw_total.columns]

# Separate 'Total' row
total_row = summary_hw_total[summary_hw_total['Circle'] == 'Total']
summary_hw_total = summary_hw_total[summary_hw_total['Circle'] != 'Total']

# Sort by 'Circle', 'OEM', and 'Total' column in descending order
summary_hw_total = summary_hw_total.sort_values(
    by=['Total'], 
    ascending=[False]
)

# Append 'Total' row at the bottom
summary_hw_total = pd.concat([summary_hw_total, total_row], ignore_index=True)
summary_hw_total = summary_hw_total.drop(columns=['Total'])



 # Create pivot table for Unique Sites count
summary_hw_unique = pd.pivot_table(
        latest7daysdump,  # DataFrame containing the data
        values='SITEID',  # Column to count
        index=['Circle','OEM'],  # Row index (Circle)+
        columns='EMS Start Date',  # Columns (EMS Start Date)
        aggfunc=lambda x: x.nunique(),  # Aggregate function (count occurrences)
        fill_value = 0,  # Fill missing values with 0       
        margins=True,  # Add totals for both rows and columns
        margins_name='Total'  # Name for the row and column totals
        
    ).reset_index()

# Format the column headers to show date in the desired format
summary_hw_unique.columns = [col.strftime('%d %b') if isinstance(col, pd.Timestamp) else col for col in summary_hw_unique.columns]

# Separate 'Total' row
total_row = summary_hw_unique[summary_hw_unique['Circle'] == 'Total']
summary_hw_unique = summary_hw_unique[summary_hw_unique['Circle'] != 'Total']

# Sort by 'Circle', 'OEM', and 'Total' column in descending order
summary_hw_unique = summary_hw_unique.sort_values(
    by=['Total'], 
    ascending=[False]
)

# Append 'Total' row at the bottom
summary_hw_unique = pd.concat([summary_hw_unique, total_row], ignore_index=True)
summary_hw_unique = summary_hw_unique.drop(columns=['Total'])


 # Create pivot table for Unique Sites count
summary_grth3daycons = pd.pivot_table(
        grth3dayscons,  # DataFrame containing the data
        values='SITEID',  # Column to count
        index=['Circle','OEM'],  # Row index (Circle)+
        columns='EMS Start Date',  # Columns (EMS Start Date)
        aggfunc=lambda x: x.nunique(),  # Aggregate function (count occurrences)
        fill_value = 0,  # Fill missing values with 0       
        margins=True,  # Add totals for both rows and columns
        margins_name='Total'  # Name for the row and column totals
        
    ).reset_index()

# Format the column headers to show date in the desired format
summary_grth3daycons.columns = [col.strftime('%d %b') if isinstance(col, pd.Timestamp) else col for col in summary_grth3daycons.columns]

# Separate 'Total' row
total_row = summary_grth3daycons[summary_grth3daycons['Circle'] == 'Total']
summary_grth3daycons = summary_grth3daycons[summary_grth3daycons['Circle'] != 'Total']

# Sort by 'Circle', 'OEM', and 'Total' column in descending order
summary_grth3daycons = summary_grth3daycons.sort_values(
    by=['Total'], 
    ascending=[False]
)

# Append 'Total' row at the bottom
summary_grth3daycons = pd.concat([summary_grth3daycons, total_row], ignore_index=True)
summary_grth3daycons = summary_grth3daycons.drop(columns=['Total'])


#=========================================================hardware Category=====================================================
df_cpri_sfp_req = Sitewise_backup[
    (Sitewise_backup['Category'].str.strip() == "CPRI/SFP/Consumable") & 
    (Sitewise_backup['Site Action Priority'].str.strip() != "Normal")
]

 # Create pivot table for Unique Sites count
summary_df_cpri_sfp_req = pd.pivot_table(
        df_cpri_sfp_req,  # DataFrame containing the data        
        index=['Circle'],  # Row index (Circle)+
        columns=['Site Action Priority'],  # Columns (EMS Start Date)
        values='SITEID',  # Column to count
        aggfunc=lambda x: x.nunique(),  # Aggregate function (count occurrences)
        fill_value = 0,  # Fill missing values with 0       
        margins=True,  # Add totals for both rows and columns
        margins_name='Total'  # Name for the row and column totals
        
    ).reset_index()

summary_df_cpri_sfp_req.rename(columns={'Total': 'CPRI/SFP/Consumable Total'}, inplace=True)

df_gps_req = Sitewise_backup[
    (Sitewise_backup['Category'].str.strip() == "GPS") & 
    (Sitewise_backup['Site Action Priority'].str.strip() != "Normal")
]

 # Create pivot table for Unique Sites count
summary_df_gps_req = pd.pivot_table(
        df_gps_req,  # DataFrame containing the data        
        index=['Circle'],  # Row index (Circle)+
        columns=['Site Action Priority'],  # Columns (EMS Start Date)
        values='SITEID',  # Column to count
        aggfunc=lambda x: x.nunique(),  # Aggregate function (count occurrences)
        fill_value = 0,  # Fill missing values with 0       
        margins=True,  # Add totals for both rows and columns
        margins_name='Total'  # Name for the row and column totals
        
    ).reset_index()

summary_df_gps_req.rename(columns={'Total': 'GPS Total'}, inplace=True)


df_hardware_req = Sitewise_backup[
    (Sitewise_backup['Category'].str.strip() == "Hardware fault") & 
    (Sitewise_backup['Site Action Priority'].str.strip() != "Normal")
]

 # Create pivot table for Unique Sites count
summary_df_hardware_req = pd.pivot_table(
        df_hardware_req,  # DataFrame containing the data        
        index=['Circle'],  # Row index (Circle)+
        columns=['Site Action Priority'],  # Columns (EMS Start Date)
        values='SITEID',  # Column to count
        aggfunc=lambda x: x.nunique(),  # Aggregate function (count occurrences)
        fill_value = 0,  # Fill missing values with 0       
        margins=True,  # Add totals for both rows and columns
        margins_name='Total'  # Name for the row and column totals
        
    ).reset_index()

summary_df_hardware_req.rename(columns={'Total': 'Hardware fault Total'}, inplace=True)


# Filter where the unique EMS Start Date count is at least 3
filtered_sites = Sitewise_backup[Sitewise_backup['Priority'] == 'P1Site(>=3/7days Alarm)']
circle_counts_P1sites = filtered_sites.groupby(['Circle'])['SITEID'].nunique().reset_index(name='P1Site(>=3/7days Alarm)')
# Calculate total
total = circle_counts_P1sites['P1Site(>=3/7days Alarm)'].sum()

# Create a DataFrame for the total row
total_row = pd.DataFrame({
    'Circle': ['Total'],
    'P1Site(>=3/7days Alarm)': [total]
})

# Append the total row
circle_counts_P1sites = pd.concat([circle_counts_P1sites, total_row], ignore_index=True)


greatertha20alarm = Sitewise_backup[Sitewise_backup['>=20 Alarms in latest Two Days'] == "P1Site(>=20 Alarm)"]
greatertha20alarm_grouped = greatertha20alarm.groupby(['Circle'])['SITEID'].nunique().reset_index(name='>= 20 Alarms')
# Calculate total
total = greatertha20alarm_grouped['>= 20 Alarms'].sum()

# Create a DataFrame for the total row
total_row = pd.DataFrame({
    'Circle': ['Total'],
    '>= 20 Alarms': [total]
})

# Append the total row
greatertha20alarm_grouped = pd.concat([greatertha20alarm_grouped, total_row], ignore_index=True)



#==================================================================Microwave Alarm=========================================================================================
# #=============================================================================Reading hardware Alarm Files============================================================
# Folder path containing the files
folder_path_mw = "D:/Microwave Hardware daily report"


# Get all .xlsx files and sort by filename (reverse for latest files first)
file_paths = glob.glob(os.path.join(folder_path_mw, "*.xlsx"))
file_paths.sort(reverse=True)  # Alphabetically descending (latest file names first)

# Get only the latest 9 files
latest_files = file_paths[:17]

# Columns to load from each file
columns_to_load = ['Circle', 'Site ID', 'Alarm Name', 'Node',
                   'First Occurrence', 'Clear Last Occurrence', 'Status',"Node Type", 'Alarm Type', 'Vendor']

# Initialize a list to hold data from each file
all_data = []

# Read each file and append the data to the list
for file_path in latest_files:
    try:
        file_size = os.path.getsize(file_path) / (1024 * 1024)  # Convert bytes to MB
        print(f"Reading file: {os.path.basename(file_path)} | Size: {file_size:.2f} MB")
        
        data = pd.read_excel(file_path, usecols=columns_to_load)
        all_data.append(data)
    except Exception as e:
        print(f"Error reading {file_path}: {e}")

# Concatenate all data into a single DataFrame
if all_data:
    combined_data_mw = pd.concat(all_data, ignore_index=True)
    print("Data successfully combined.")
else:
    combined_data_mw = pd.DataFrame(columns=columns_to_load)
    print("No valid data found.")

#===========================================================================processing High Temp Fan========================================================================
combined_data_mw = combined_data_mw.dropna(subset=['First Occurrence'])
combined_data_mw.rename(columns={'Alarm Name': 'Alarm_Name', "Vendor":"OEM", "Site ID": "SITEID"}, inplace=True)
#combined_data_mw = combined_data_mw.dropna(subset=['EMS Start Time'])
combined_data_mw['EMS Start Date'] = combined_data_mw['First Occurrence'].dt.date
combined_data_mw['EMS Start Date'] = pd.to_datetime(combined_data_mw['EMS Start Date'], errors='coerce')
# Get today's datee
today_date = pd.to_datetime('today').normalize()
# Filter out rows where 'EMS Start Date' is today
combined_data_mw = combined_data_mw[combined_data_mw['EMS Start Date'] != today_date]
combined_data_mw = combined_data_mw[
    ~combined_data_mw['SITEID'].str.contains("Unknown", case=False, na=False)
]
combined_data_mw['Week Number'] = combined_data_mw['EMS Start Date'].dt.isocalendar().week
combined_data_mw['Week Name'] = 'WK-' + combined_data_mw['EMS Start Date'].dt.isocalendar().week.astype(str)

# Extract unique dates and drop NaT values
unique_dates = combined_data_mw['EMS Start Date'].unique()
# Sort the unique dates and format them
readable_dates = [date.strftime('%d-%m-%Y') for date in sorted(unique_dates)]
# Print the unique, readable dates
print("Unique EMS Start Dates:")
for date in readable_dates:
    print(date)

#start_date_input = "23-01-2025"

try:
    # Convert start date to datetime
    #start_date = pd.to_datetime(start_date_input, format='%d-%m-%Y', errors='coerce')

    # Ensure 'EMS Start Date' is in datetime format
    combined_data_mw['EMS Start Date'] = pd.to_datetime(combined_data_mw['EMS Start Date'], errors='coerce')

    # Get max EMS Start Date
    
    end_date = combined_data_mw['EMS Start Date'].max()
    start_date = end_date - pd.Timedelta(days=15)

    # Check if dates are valid
    if pd.isna(start_date) or pd.isna(end_date):
        print("Invalid date format or missing data.")
    elif start_date > end_date:
        print("Start date cannot be later than the max EMS Start Date.")
    else:
        # Filter data
        filtered_data_mw = combined_data_mw[
            (combined_data_mw['EMS Start Date'] >= start_date) &
            (combined_data_mw['EMS Start Date'] <= end_date)
        ]

        print(f"Filtered data between {start_date.strftime('%d-%m-%Y')} and {end_date.strftime('%d-%m-%Y')}:")

except Exception as e:
    print(f"An error occurred: {e}")

filtered_data_mw['Date-Month'] = filtered_data_mw['EMS Start Date'].dt.strftime('%d-%m')
#==========================================================calculating consecutive days=======================================

def calculate_alarm_period(group):
    group = group.sort_values('EMS Start Date', ascending=True)  # Sort ascending order
    
    # Initialize Consecutive_Days column
    group['Consecutive_Days'] = 1  
    
    # Iterate through rows to manually calculate consecutive days
    for i in range(1, len(group)):
        if group.iloc[i]['EMS Start Date'] == group.iloc[i - 1]['EMS Start Date']:
            # If EMS Start Date is same as previous, keep the same Consecutive_Days
            group.loc[group.index[i], 'Consecutive_Days'] = group.loc[group.index[i - 1], 'Consecutive_Days']
        elif (group.iloc[i]['EMS Start Date'] - group.iloc[i - 1]['EMS Start Date']).days == 1:
            # If the date is exactly +1 day, increase Consecutive_Days
            group.loc[group.index[i], 'Consecutive_Days'] = group.loc[group.index[i - 1], 'Consecutive_Days'] + 1
        else:
            # If there is a gap (missing date), reset to 1
            group.loc[group.index[i], 'Consecutive_Days'] = 1

    return group

filtered_data_mw = filtered_data_mw.groupby(
    ['Circle', 'OEM', 'SITEID', 'Alarm_Name', "Node Type",'Alarm Type'],
    group_keys=False
).apply(calculate_alarm_period)

#===========================================================Defining Alarm Occurance Period=================================================

# Assign 'Alarm Occurrence Period' where Consecutive Days are 3 or more
filtered_data_mw['Alarm Occurrence Period'] = np.where(filtered_data_mw['Consecutive_Days'] >= 3, '≥ 3 Days', 'Normal')

# Final sorting to maintain order
filtered_data_mw = filtered_data_mw.sort_values(['SITEID', 'EMS Start Date'], ascending=[True, True])


latest_date = filtered_data_mw['EMS Start Date'].max()

#Filter data for the latest 3 days
latest1daysdump_mw = filtered_data_mw[filtered_data_mw['EMS Start Date'] > (latest_date - pd.Timedelta(days=1))]
latest2daysdump_mw = filtered_data_mw[filtered_data_mw['EMS Start Date'] >= (latest_date - pd.Timedelta(days=1))]
latest3daysdump_mw = filtered_data_mw[filtered_data_mw['EMS Start Date'] >= (latest_date - pd.Timedelta(days=2))]
latest7daysdump_mw = filtered_data_mw[filtered_data_mw['EMS Start Date'] >= (latest_date - pd.Timedelta(days=6))]
grth3dayscons_mw = latest7daysdump_mw[latest7daysdump_mw['Alarm Occurrence Period'] == '≥ 3 Days']


#======================================================Working on Sitewise_Backup_Daily==================================================================================

# Count unique EMS Start Dates per Circle and SITEID
site_date_counts_mw = latest7daysdump_mw.groupby(['Circle', 'OEM', 'SITEID', 'Alarm_Name', "Node Type",'Alarm Type'])['EMS Start Date'].nunique().reset_index(name='Count in Days/7 Days')

# Filter where the unique EMS Start Date count is at least 3
filtered_sites_mw = site_date_counts_mw[site_date_counts_mw['Count in Days/7 Days'] >= 3]
# Count how many such SITEIDs exist per Circle
circle_counts_mw = filtered_sites_mw.groupby(['Circle'])['SITEID'].nunique().reset_index(name='P1Site(>=3/7days Alarm)')


# Create Pivot Table
Sitewise_backup_mw = filtered_data_mw.pivot_table(
    index=['Circle', 'OEM', 'SITEID', 'Alarm_Name', "Node Type",'Alarm Type'], 
    columns='EMS Start Date', 
    values='Date-Month',  # Use any column that exists for counting
    aggfunc= 'count',
    fill_value=np.nan,  # Fill missing values with 0
    margins=True,  # Add totals for both rows
    margins_name='Total'  # Name for the row and column totals
).reset_index()

# Format the column headers to show date in the desired format
Sitewise_backup_mw.columns = [col.strftime('%d %b') if isinstance(col, pd.Timestamp) else col for col in Sitewise_backup_mw.columns]
Sitewise_backup_mw = Sitewise_backup_mw[Sitewise_backup_mw['Circle'] != 'Total']


Sitewise_backup_mw = Sitewise_backup_mw.merge(site_date_counts_mw, on=['Circle', 'OEM', 'SITEID', 'Alarm_Name', "Node Type",'Alarm Type'], how="left")

# Assuming Sitewise_backup and filtered_sites are already defined DataFrames
Sitewise_backup_mw["Priority"] = Sitewise_backup_mw["Count in Days/7 Days"].apply(
    lambda x: "P1Site(>=3/7days Alarm)" if x >= 3 else "Normal"
)


# Define conditions and corresponding values
conditions = [
    Sitewise_backup_mw["Count in Days/7 Days"].isin([1, 2]),  # If value is 1 or 2
    Sitewise_backup_mw["Count in Days/7 Days"].isin([3, 4]),  # If value is 3 or 4
    Sitewise_backup_mw["Count in Days/7 Days"] >= 5           # If value is 5 or more
]

values = ["1-2 days/7days", "3-4 days/7days", ">=5 days/7days"]

# Apply the classification
Sitewise_backup_mw["Site Action Priority"] = np.select(conditions, values, default="Normal")


#=============================================>= 20 Alarms in two consecutive days and >= 3 Days Cosecutive=============================================

ems_dates = list(Sitewise_backup_mw.columns[20:])
Sitewise_backup_mw['>=20 Alarms in latest Two Days'] = np.where(
    (Sitewise_backup_mw[ems_dates[0]] >= 20) & (Sitewise_backup_mw[ems_dates[1]] >= 20), 
    "P1Site(>=20 Alarm)", "Normal"
)

ems_dates = list(Sitewise_backup_mw.columns[19:])
Sitewise_backup_mw['>= 3 Days Cosecutive'] = np.where(
    (Sitewise_backup_mw[ems_dates[0]] > 0) & (Sitewise_backup_mw[ems_dates[1]] > 0) & (Sitewise_backup_mw[ems_dates[2]] > 0), 
    ">= 3 Days", "Normal"
)





greatertha20alarm_mw = Sitewise_backup_mw[Sitewise_backup_mw['>=20 Alarms in latest Two Days'] == "P1Site(>=20 Alarm)"]
greatertha20alarm_grouped_mw = greatertha20alarm_mw.groupby(['Circle'])['SITEID'].nunique().reset_index(name='>= 20 Alarms MW')
# Calculate total
total = greatertha20alarm_grouped_mw['>= 20 Alarms MW'].sum()

# Create a DataFrame for the total row
total_row = pd.DataFrame({
    'Circle': ['Total'],
    '>= 20 Alarms MW': [total]
})

# Append the total row
greatertha20alarm_grouped_mw = pd.concat([greatertha20alarm_grouped_mw, total_row], ignore_index=True)





# Filter where the unique EMS Start Date count is at least 3
filtered_sites_mw = Sitewise_backup_mw[Sitewise_backup_mw['Priority'] == 'P1Site(>=3/7days Alarm)']
circle_counts_P1sites_mw = filtered_sites_mw.groupby(['Circle'])['SITEID'].nunique().reset_index(name='P1Site(>=3/7days Alarm) MW')
# Calculate total
total = circle_counts_P1sites_mw['P1Site(>=3/7days Alarm) MW'].sum()

# Create a DataFrame for the total row
total_row = pd.DataFrame({
    'Circle': ['Total'],
    'P1Site(>=3/7days Alarm) MW': [total]
})

# Append the total row
circle_counts_P1sites_mw = pd.concat([circle_counts_P1sites_mw, total_row], ignore_index=True)





#==================================================Hardware_Summary based on last 7 days data=============================================

summary_hw_total_mw = latest7daysdump_mw.pivot_table(
    index=['Circle', 'OEM'], 
    columns='EMS Start Date', 
    values='SITEID',  # Use any column that exists for counting
    aggfunc= 'count',
    fill_value = 0,  # Fill missing values with 0
    margins=True,  # Add totals for both rows
    margins_name='Total'  # Name for the row and column totals
).reset_index()

# Format the column headers to show date in the desired format
summary_hw_total_mw.columns = [col.strftime('%d %b') if isinstance(col, pd.Timestamp) else col for col in summary_hw_total_mw.columns]

# Separate 'Total' row
total_row = summary_hw_total_mw[summary_hw_total_mw['Circle'] == 'Total']
summary_hw_total_mw = summary_hw_total_mw[summary_hw_total_mw['Circle'] != 'Total']

# Sort by 'Circle', 'OEM', and 'Total' column in descending order
summary_hw_total_mw = summary_hw_total_mw.sort_values(
    by=['Total'], 
    ascending=[False]
)

# Append 'Total' row at the bottom
summary_hw_total_mw = pd.concat([summary_hw_total_mw, total_row], ignore_index=True)
summary_hw_total_mw = summary_hw_total_mw.drop(columns=['Total'])



 # Create pivot table for Unique Sites count
summary_hw_unique_mw = pd.pivot_table(
        latest7daysdump_mw,  # DataFrame containing the data
        values='SITEID',  # Column to count
        index=['Circle','OEM'],  # Row index (Circle)+
        columns='EMS Start Date',  # Columns (EMS Start Date)
        aggfunc=lambda x: x.nunique(),  # Aggregate function (count occurrences)
        fill_value = 0,  # Fill missing values with 0       
        margins=True,  # Add totals for both rows and columns
        margins_name='Total'  # Name for the row and column totals
        
    ).reset_index()

# Format the column headers to show date in the desired format
summary_hw_unique_mw.columns = [col.strftime('%d %b') if isinstance(col, pd.Timestamp) else col for col in summary_hw_unique_mw.columns]

# Separate 'Total' row
total_row = summary_hw_unique_mw[summary_hw_unique_mw['Circle'] == 'Total']
summary_hw_unique_mw = summary_hw_unique_mw[summary_hw_unique_mw['Circle'] != 'Total']

# Sort by 'Circle', 'OEM', and 'Total' column in descending order
summary_hw_unique_mw = summary_hw_unique_mw.sort_values(
    by=['Total'], 
    ascending=[False]
)

# Append 'Total' row at the bottom
summary_hw_unique_mw = pd.concat([summary_hw_unique_mw, total_row], ignore_index=True)
summary_hw_unique_mw = summary_hw_unique_mw.drop(columns=['Total'])


 # Create pivot table for Unique Sites count
summary_grth3daycons_mw = pd.pivot_table(
        grth3dayscons_mw,  # DataFrame containing the data
        values='SITEID',  # Column to count
        index=['Circle','OEM'],  # Row index (Circle)+
        columns='EMS Start Date',  # Columns (EMS Start Date)
        aggfunc=lambda x: x.nunique(),  # Aggregate function (count occurrences)
        fill_value = 0,  # Fill missing values with 0       
        margins=True,  # Add totals for both rows and columns
        margins_name='Total'  # Name for the row and column totals
        
    ).reset_index()

# Format the column headers to show date in the desired format
summary_grth3daycons_mw.columns = [col.strftime('%d %b') if isinstance(col, pd.Timestamp) else col for col in summary_grth3daycons_mw.columns]

# Separate 'Total' row
total_row = summary_grth3daycons_mw[summary_grth3daycons_mw['Circle'] == 'Total']
summary_grth3daycons_mw = summary_grth3daycons_mw[summary_grth3daycons_mw['Circle'] != 'Total']

# Sort by 'Circle', 'OEM', and 'Total' column in descending order
summary_grth3daycons_mw = summary_grth3daycons_mw.sort_values(
    by=['Total'], 
    ascending=[False]
)

# Append 'Total' row at the bottom
summary_grth3daycons_mw = pd.concat([summary_grth3daycons_mw, total_row], ignore_index=True)
summary_grth3daycons_mw = summary_grth3daycons_mw.drop(columns=['Total'])
#==================================================================Category Summary===================================================
# Merge the DataFrames using left join
summary_req_all = (
    summary_df_cpri_sfp_req
    .merge(summary_df_gps_req, on='Circle', how='left')
    .merge(summary_df_hardware_req, on='Circle', how='left')
    .merge(greatertha20alarm_grouped, on='Circle', how='left')
    .merge(circle_counts_P1sites, on='Circle', how='left')
    .merge(greatertha20alarm_grouped_mw, on='Circle', how='left')
    .merge(circle_counts_P1sites_mw, on='Circle', how='left')
    .fillna(0)  # Fill NaNs with 0 after merging
)
#=================================================================Dashboard High Temp=======================================================================================

report_date = combined_data_high_temp['EMS Start Date'].max().strftime('%d-%m-%Y')
output_folder = r"D:\\Automation\\Hardware Alarms\\Output"
#yesterday = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime("%Y%m%d")
output_file = os.path.join(output_folder, f'RAN and MW Hardware Alarms_{report_date}.xlsx')


# Create output folder if it doesn't exist
os.makedirs(output_folder, exist_ok=True)


# Write to Excel
with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:     
    
    summary_req_all.to_excel(writer, sheet_name="Summary", index=False,startrow = 1)
    summary_hw_total.to_excel(writer, sheet_name="Summary", index=False,startcol = 0,startrow = 30)
    summary_hw_unique.to_excel(writer, sheet_name="Summary", index=False,startcol = 10,startrow = 30)
    summary_grth3daycons.to_excel(writer, sheet_name="Summary", index=False,startcol = 20,startrow = 30)
    summary_hw_total_mw.to_excel(writer, sheet_name="Summary", index=False,startcol = 0,startrow = 62)
    summary_hw_unique_mw.to_excel(writer, sheet_name="Summary", index=False,startcol = 10,startrow = 62)
    summary_grth3daycons_mw.to_excel(writer, sheet_name="Summary", index=False,startcol = 20,startrow = 62)
    Sitewise_backup.to_excel(writer, sheet_name="Sitewise_Backup_Daily", index=False) 
    Sitewise_backup_mw.to_excel(writer, sheet_name="Sitewise_Backup_Daily_MW", index=False)
    latest3daysdump.to_excel(writer, sheet_name="Dump_Latest_3_Days", index=False)


   

    # Apply formatting
    workbook  = writer.book
    worksheet = writer.sheets["Summary"]

    merge_format = workbook.add_format({
        'bold': True,
        'align': 'center',
        'valign': 'vcenter',
        'border': 1
    })

    # Use the last row (usually the "Total" row)
    last_row = summary_req_all.iloc[-1]

    cpri_fault = last_row[4]
    GPS_Fault = last_row[8]
    hardware_Fault = last_row[12]
    grth20alarms = last_row[13]
    p1sites = last_row[14]
    grth20alarmsmw = last_row[15]
    p1sitesmw = last_row[16]
    # Merge cells with labels

    worksheet.merge_range(0, 1, 0, 4, 'CPRI/SFP/Consumable', merge_format)
    worksheet.merge_range(0, 5, 0, 8, 'GPS', merge_format)
    worksheet.merge_range(0, 9, 0, 12, 'Hardware fault', merge_format)
    worksheet.merge_range(0, 13, 0, 14, 'Alarm Type RAN', merge_format)
    worksheet.merge_range(0, 15, 0, 16, 'Alarm Type MW', merge_format)


    worksheet.merge_range(29, 0, 29, 8, 'Total Sites Hardware Events', merge_format)
    worksheet.merge_range(29, 10, 29, 18, 'Unique Sites Hardware Events', merge_format)
    worksheet.merge_range(29, 20, 29, 28, '>= 3 Consecutive Days Site Count', merge_format)

    worksheet.merge_range(61, 0, 61, 8, 'Total Sites Hardware Events (MW)', merge_format)
    worksheet.merge_range(61, 10, 61, 18, 'Unique Sites Hardware Events (MW)', merge_format)
    worksheet.merge_range(61, 20, 61, 28, '>= 3 Consecutive Days Site Count(MW)', merge_format)
    


print("Data saved successfully to multiple sheets in:",output_file)


#===============================================================================================

# # Save the data to Excel
# #current_date = datetime.datetime.now().strftime("%Y%m%d")
# output_file_path = rf"D:\\Automation\\Hardware Alarms\\Output\\RAN Hardware Alarms.xlsx"

# try:
#     book = load_workbook(output_file_path)

#     # Sheets to delete if they exist
#     sheets_to_remove = ["Sitewise_Backup_Daily", "Dump_Latest_3_Days", "Sitewise_Backup_Daily_MW"]

#     for sheet_name in sheets_to_remove:
#         if sheet_name in book.sheetnames:
#             std = book[sheet_name]
#             book.remove(std)

#     book.save(output_file_path)  # Save after removal

#     # Re-open writer for appending with updated workbook
#     writer = pd.ExcelWriter(output_file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay")
#     writer._book = book  # Proper assignment for openpyxl writer

# except FileNotFoundError:
#     writer = pd.ExcelWriter(output_file_path, engine="openpyxl", mode="w")

# #greatertha50alarm_grouped .to_excel(writer, sheet_name="Hardware_Summary", index=False,startcol = 1,startrow = 3)
# summary_req_all.to_excel(writer, sheet_name="Summary", index=False,startrow = 1)

# summary_hw_total.to_excel(writer, sheet_name="Summary", index=False,startcol = 0,startrow = 30)
# summary_hw_unique.to_excel(writer, sheet_name="Summary", index=False,startcol = 10,startrow = 30)
# summary_grth3daycons.to_excel(writer, sheet_name="Summary", index=False,startcol = 20,startrow = 30)


# summary_hw_total_mw.to_excel(writer, sheet_name="Summary", index=False,startcol = 0,startrow = 62)
# summary_hw_unique_mw.to_excel(writer, sheet_name="Summary", index=False,startcol = 10,startrow = 62)
# summary_grth3daycons_mw.to_excel(writer, sheet_name="Summary", index=False,startcol = 20,startrow = 62)


# Sitewise_backup.to_excel(writer, sheet_name="Sitewise_Backup_Daily", index=False) 
# Sitewise_backup_mw.to_excel(writer, sheet_name="Sitewise_Backup_Daily_MW", index=False)
# latest3daysdump.to_excel(writer, sheet_name="Dump_Latest_3_Days", index=False)
# #greatertha50alarm.to_excel(writer, sheet_name="Greater Than 50 Alarm", index=False)
# #summary_category.to_excel(writer, sheet_name="site_priority_Category", index=True)
    
# writer.close()



# # Define thick border
# thin_border = Border(
#     left=Side(style='thin'),
#     right=Side(style='thin'),
#     top=Side(style='thin'),
#     bottom=Side(style='thin')
# )

# # Define red background and white bold text
# fill = PatternFill(start_color="FF0000", fill_type="solid")  # Red background
# font_white = Font(color="FFFFFF", bold=True)  # White bold text
# wrap_middle_alignment = Alignment(wrap_text=True, vertical='center') 

# for col in list(range(1, 18)):
#     cell = ws.cell(row=2, column=col)
#     cell.alignment = wrap_middle_alignment

# for col in  list(range(1, 10)) + list(range(11, 20)) + list(range(21, 30)):
#     cell = ws.cell(row=31, column=col)
#     cell.font = font_white

# for col in  list(range(1, 10)) + list(range(11, 20)) + list(range(21, 30)):
#     cell = ws.cell(row=63, column=col)
#     cell.font = font_white
  



#book.save(output_file)  # Save the final formatting

#=================================sending Mail=====================================

# Read recipients from Excel
file_path_receipient = r"D:\Automation\Hardware Alarms\Receipient.xlsx"
try:
    receipient_df = pd.read_excel(file_path_receipient)
    to_recipients = receipient_df["to"].dropna().tolist()  # List of 'To' recipients
    cc_recipients = receipient_df["cc"].dropna().tolist()  # List of 'CC' recipients
except Exception as e:
    print(f"❌ Error reading Excel file: {e}")
    sys.exit()

# Define sender email
from_email = "IN_D_Engineering_NIfi_Reports@airtel.com"

# Initialize Outlook
try:
    outlook = win32com.client.Dispatch("Outlook.Application")
    namespace = outlook.GetNamespace("MAPI")
except Exception as e:
    print(f"❌ Error initializing Outlook: {e}")
    sys.exit()

# Get today's date for email subject
today_date = datetime.now().strftime("%d%m%Y")
yesterday_date = (datetime.now() - timedelta(days=1)).strftime("%d%m%Y")



email_subject = f"RAN/MW Hardware Alarms Summary - {yesterday_date}"
email_body = f"""
Dear Team,

Please find RAN and MW Hardware fault report is shared daily for NOK,Eric,SAMSUNG and ZTE. Still many circles are not sharing daily update on the progress.
Pls address the repeat fault sites on priority to reduce the active outages.
>=5/7 days
CPRI/Consumable/SFP faults: {cpri_fault}
GPS: {GPS_Fault} ( JK is due to interference issue)
Hardware faults : {hardware_Fault}
>=20 Alarm(last 2 days rolling): {grth20alarms}
P1Site(>=3/7days Alarm): {p1sites}
>=20 Alarm(last 2 days rolling) MW: {grth20alarmsmw}
P1Site(>=3/7days Alarm) MW: {p1sitesmw}


Regards,  
Central Team FSO
"""

#output_file_path = r"D:\Automation\Hardware Alarms\Output\RAN Hardware Alarms.xlsx"  # Adjust path

try:
    # Create a new email
    mail = outlook.CreateItem(0)
    mail.To = ";".join(to_recipients) if to_recipients else ""  # Join multiple recipients with ";"
    mail.CC = ";".join(cc_recipients) if cc_recipients else ""  # Join multiple CC recipients
    mail.Subject = email_subject
    mail.Body = email_body

    # Specify the sender email
    mail.SentOnBehalfOfName = from_email  # Option 1 (if account has permission)

    # Alternative approach (set the sender account explicitly)
    for account in outlook.Session.Accounts:
        if account.SmtpAddress == from_email:
            mail._oleobj_.Invoke(*(64209, 0, 8, 0, account))  # Option 2 (SendUsingAccount)

    # Attach the file
    mail.Attachments.Add(output_file)
    mail.Send()  # Send the email

    print(f"📧 Email sent successfully from {from_email}")

except Exception as e:
    print(f"❌ Error sending email: {e}")

# Release Outlook objects
del outlook, namespace

print("✅ Email process completed successfully!")   