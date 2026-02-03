import pandas as pd
import numpy as np
import os
import datetime as dt
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import (Alignment, PatternFill, Font, Border, Side, numbers)
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
# ==================================== Functions ==================================== #

def browse_vehicle_master():
    """Allow the user to select the Vehicle Master file."""
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    vehicle_master_entry.delete(0, tk.END)
    vehicle_master_entry.insert(0, file_path)

def browse_monthly_dump():
    """Allow the user to select the Monthly Dump CSV file."""
    file_path = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
    if file_path:  # Ensure the user selected a file
        monthly_dump_entry.delete(0, tk.END)
        monthly_dump_entry.insert(0, file_path)  # Show the full file path


def browse_output_folder():
    """Allow the user to select the Output folder."""
    folder_path = filedialog.askdirectory()
    output_folder_entry.delete(0, tk.END)
    output_folder_entry.insert(0, folder_path)

def process_files():
    """Process the selected files and generate the output."""
    vehicle_master_path = vehicle_master_entry.get()
    monthly_dump_folder = monthly_dump_entry.get()
    output_folder = output_folder_entry.get()

    if not vehicle_master_path or not monthly_dump_folder or not output_folder:
        messagebox.showerror("Error", "Please select all required paths.")
        return

    try:
        # ================ Reading Vehicle Master Data ================ #
        vehicle_master = pd.read_excel(vehicle_master_path, sheet_name='Vehicle')
        vehicle_master = vehicle_master.drop(columns=['Supplier Name', 'Model'])

        replace_dict = {
            'AP-TL': 'AP', 'NESA - AS': 'AS', 'NESA - NE': 'NE', 'BH - BH': 'BR', 'BH - JH': 'BR',
            'KTN - CHN': 'CN', 'KTN - TN': 'TN', 'HPHP - HP': 'HP', 'HPHP - HR': 'HR', 'KTN - KL': 'KL',
            'WB - KOL': 'KO', 'MH/GOA': 'MH', 'MPCG - MP': 'MP', 'MPCG - CG': 'MP', 'HPHP - PB/CHD': 'PB',
            'UPE': "UE", 'UPW': "UW", 'UPW-UT': "UW", 'WB - ROB': 'WB'
        }
        vehicle_master['Circle'] = vehicle_master['Circle'].replace(replace_dict, regex=False)
        vehicle_master = vehicle_master.groupby("Circle", as_index=False)['Vehicles'].sum()
        vehicle_master = vehicle_master.rename(columns={'Vehicles': 'Total Vehicles in Circle'})
        # Add a "Total" row
        total_vehicles = vehicle_master['Total Vehicles in Circle'].sum()
        vehicle_master.loc[len(vehicle_master)] = ['Total', total_vehicles]

        # ================ Reading Monthly Dump Data ================ #
        columns_to_load = [
            "date", "circle", "trip_id", "vehicle_id", "vehicle_no", "vehicle_agency",
            "driver_name", "driver_number", "first_order_raised", "driver_start_time",
            "first_fme_reached_time", "fme_day_last_end_time", "driver_day_last_end_time",
            "no_of_trips", "no_of_users", "total_wos_handled_with_vehicle", "total_km",
            "garage_km", "total_night_hours", "user_name_1", "user_olm_id_1", "user_msisdn_1",
            "user_1_start_time", "user_1_end_time", "user_1_km", "user_1_travel_time",
            "user_1_work_orders", "user_1_no_of_orders", "user_1_unique_sites"
        ]

        monthly_dump_path = monthly_dump_entry.get()
        daily_vehicle_dump = pd.read_csv(monthly_dump_path, usecols=columns_to_load)

    
        daily_vehicle_dump.rename(columns={'circle': 'Circle'}, inplace=True)
        date_cols = [
            'date', 'first_order_raised', 'driver_start_time', 'first_fme_reached_time',
            'fme_day_last_end_time', 'driver_day_last_end_time', 'user_1_start_time', 'user_1_end_time'
        ]
        for col in date_cols:
            daily_vehicle_dump[col] = pd.to_datetime(daily_vehicle_dump[col], errors='coerce')

        # ================ Feature Engineering ================ #
        daily_vehicle_dump['Vehicle Used'] = np.where(daily_vehicle_dump['driver_start_time'].notna(), 'Vehicle Reported', 'Vehicle Not Reported')
        daily_vehicle_dump['Vehicle Onboarded'] = np.where(daily_vehicle_dump['first_fme_reached_time'].notna(), 'FME Used Vehicle', 'FME Not Used Vehicle')
        daily_vehicle_dump['WO Handling with Vehicle'] = np.where(daily_vehicle_dump['total_wos_handled_with_vehicle'].notna(), 'WO Handled with Vehicle', 'No WO Handled with Vehicle')
        daily_vehicle_dump['Vehicles Requested after 10AM by FME'] = daily_vehicle_dump['first_order_raised'].dt.hour
        daily_vehicle_dump['Vehicle Reporting Time'] = daily_vehicle_dump['first_fme_reached_time'].dt.hour
        daily_vehicle_dump['Vehicle release time'] = daily_vehicle_dump['fme_day_last_end_time'].dt.hour
        daily_vehicle_dump['Vehicle Used Time'] = ((daily_vehicle_dump['fme_day_last_end_time'] - daily_vehicle_dump['first_fme_reached_time']).dt.total_seconds() / 3600).round(2).fillna(0).astype(float)

        conditions = [
            (daily_vehicle_dump['Vehicle Used Time'] < 0.1),
            (daily_vehicle_dump['Vehicle Used Time'] < 4),
            (daily_vehicle_dump['Vehicle Used Time'] > 12)
        ]
        choices = ["Trip not closed", "Used <4 hours", "Used >12 hours"]
        daily_vehicle_dump['Vehicle Trip Time'] = np.select(conditions, choices, default="Used 4-12 hours")







        daily_vehicle_dump['Trip Close Status'] = np.select(
            [daily_vehicle_dump['first_fme_reached_time'].isna(),
             daily_vehicle_dump['driver_day_last_end_time'].isna(),
             daily_vehicle_dump['fme_day_last_end_time'].isna()],
            ["Vehicle not used", "Trip Not Closed", "Trip Not Closed"], default="Trip Closed"
        )


        # Define conditions
        conditions = [
        (daily_vehicle_dump['first_fme_reached_time'].notna()) & (daily_vehicle_dump['fme_day_last_end_time'].isna()),
        (daily_vehicle_dump['first_fme_reached_time'].notna()) & (daily_vehicle_dump['driver_day_last_end_time'].isna())
                    ]

        # Define corresponding values
        choices = [
            "TRIP Not closed by FME",
            "TRIP Not closed by Driver"
        ]

        # Apply the conditions using np.select
        daily_vehicle_dump['Trip Not Closed by'] = np.select(conditions, choices, default="TRIP Closed")
        # ======================= Dashboard ==================== #
        #vehreqfme = daily_vehicle_dump.groupby(['circle']).size().reset_index(name='Vehicles Requested by FMEs')




        vehreqfme = daily_vehicle_dump.groupby(['Circle']).size().reset_index(name='Vehicles Requested by FMEs')
        vehreqfme.loc['Total'] = ['Total', vehreqfme['Vehicles Requested by FMEs'].sum()]
        vehicle_master = vehicle_master.merge(vehreqfme, on="Circle", how="left")
        vehicle_master['Vehicles Requested by FMEs'] = vehicle_master['Vehicles Requested by FMEs'].fillna(0).astype(int)
        vehicle_master['Vehicles Not Requested by FMEs'] = vehicle_master['Total Vehicles in Circle'] - vehicle_master['Vehicles Requested by FMEs']

        dfvehreported = daily_vehicle_dump[daily_vehicle_dump['Vehicle Used'] == "Vehicle Reported"]
        dfvehreported = dfvehreported.groupby(['Circle']).size().reset_index(name='Total Vehicles Reported ( driver accepted request)')
        dfvehreported.loc['Total'] = ['Total', dfvehreported['Total Vehicles Reported ( driver accepted request)'].sum()]
        vehicle_master = vehicle_master.merge(dfvehreported, on="Circle", how="left")
        vehicle_master['Total Vehicles Reported ( driver accepted request)'] = vehicle_master['Total Vehicles Reported ( driver accepted request)'].fillna(0).astype(int)


        vehicle_master['Vehicles requested but not accepted requests'] = vehicle_master['Vehicles Requested by FMEs'] - vehicle_master['Total Vehicles Reported ( driver accepted request)']

        vehicle_master['Absent Vehicles'] = vehicle_master['Total Vehicles in Circle'] - vehicle_master['Total Vehicles Reported ( driver accepted request)']

        vehicle_master['Absent Vehicle %age'] = (
            (vehicle_master['Absent Vehicles'] / vehicle_master['Total Vehicles in Circle'])
            .fillna(0) * 100
        ).round(0).astype(int).astype(str) + '%'

        dfvehonboarded = daily_vehicle_dump[daily_vehicle_dump['Vehicle Onboarded'] == "FME Used Vehicle"]
        dfvehonboarded = dfvehonboarded.groupby(['Circle']).size().reset_index(name='Vehicles Onboarded by FMEs')
        dfvehonboarded.loc['Total'] = ['Total', dfvehonboarded['Vehicles Onboarded by FMEs'].sum()]
        vehicle_master = vehicle_master.merge(dfvehonboarded, on="Circle", how="left")
        vehicle_master['Vehicles Onboarded by FMEs'] = vehicle_master['Vehicles Onboarded by FMEs'].fillna(0).astype(int)

        dfvehwohandeled = daily_vehicle_dump[daily_vehicle_dump['WO Handling with Vehicle'] == "WO Handled with Vehicle"]
        dfvehwohandeled = dfvehwohandeled.groupby(['Circle']).size().reset_index(name='Vehicles handled WOs')
        dfvehwohandeled.loc['Total'] = ['Total', dfvehwohandeled['Vehicles handled WOs'].sum()]
        vehicle_master = vehicle_master.merge(dfvehwohandeled, on="Circle", how="left")
        vehicle_master['Vehicles handled WOs'] = vehicle_master['Vehicles handled WOs'].fillna(0).astype(int)
        vehicle_master['Vehicles onboarded but WO not Handled'] = vehicle_master['Vehicles Onboarded by FMEs'] - vehicle_master['Vehicles handled WOs']

        dfvehonewohandeled = daily_vehicle_dump[daily_vehicle_dump['total_wos_handled_with_vehicle'] == 1]
        dfvehonewohandeled = dfvehonewohandeled.groupby(['Circle']).size().reset_index(name='Vehicles Handled 1 WO')
        dfvehonewohandeled.loc['Total'] = ['Total', dfvehonewohandeled['Vehicles Handled 1 WO'].sum()]
        vehicle_master = vehicle_master.merge(dfvehonewohandeled, on="Circle", how="left")
        vehicle_master['Vehicles Handled 1 WO'] = vehicle_master['Vehicles Handled 1 WO'].fillna(0).astype(int)


        dfveh23wohandeled = daily_vehicle_dump[daily_vehicle_dump['total_wos_handled_with_vehicle'].isin([2, 3])]
        dfveh23wohandeled = dfveh23wohandeled.groupby(['Circle']).size().reset_index(name='Vehicles Handled 2-3 WO')
        dfveh23wohandeled.loc['Total'] = ['Total', dfveh23wohandeled['Vehicles Handled 2-3 WO'].sum()]
        vehicle_master = vehicle_master.merge(dfveh23wohandeled, on="Circle", how="left")
        vehicle_master['Vehicles Handled 2-3 WO'] = vehicle_master['Vehicles Handled 2-3 WO'].fillna(0).astype(int)

        dfvehgrth3wohandeled = daily_vehicle_dump[daily_vehicle_dump['total_wos_handled_with_vehicle'] > 3]
        dfvehgrth3wohandeled = dfvehgrth3wohandeled.groupby(['Circle']).size().reset_index(name='Vehicles Handled >3 WOs')
        dfvehgrth3wohandeled.loc['Total'] = ['Total', dfvehgrth3wohandeled['Vehicles Handled >3 WOs'].sum()]
        vehicle_master = vehicle_master.merge(dfvehgrth3wohandeled, on="Circle", how="left")
        vehicle_master['Vehicles Handled >3 WOs'] = vehicle_master['Vehicles Handled >3 WOs'].fillna(0).astype(int)

        dfvehrqafter10 = daily_vehicle_dump[daily_vehicle_dump['Vehicles Requested after 10AM by FME'] > 9]   #   Need to Clear
        dfvehrqafter10 = dfvehrqafter10.groupby(['Circle']).size().reset_index(name='Vehicles Requested after 10AM')
        dfvehrqafter10.loc['Total'] = ['Total', dfvehrqafter10['Vehicles Requested after 10AM'].sum()]
        vehicle_master = vehicle_master.merge(dfvehrqafter10, on="Circle", how="left")
        vehicle_master['Vehicles Requested after 10AM'] = vehicle_master['Vehicles Requested after 10AM'].fillna(0).astype(int)

        dfvehonboardafter11 = daily_vehicle_dump[daily_vehicle_dump['Vehicle Reporting Time'] > 10]   #   Need to Clear
        dfvehonboardafter11 = dfvehonboardafter11.groupby(['Circle']).size().reset_index(name='Vehicles Onboarded after 11AM')
        dfvehonboardafter11.loc['Total'] = ['Total', dfvehonboardafter11['Vehicles Onboarded after 11AM'].sum()]
        vehicle_master = vehicle_master.merge(dfvehonboardafter11, on="Circle", how="left")
        vehicle_master['Vehicles Onboarded after 11AM'] = vehicle_master['Vehicles Onboarded after 11AM'].fillna(0).astype(int)

        dfvehonboardbefore6 = daily_vehicle_dump[daily_vehicle_dump['Vehicle Reporting Time'] < 6]   #   Need to Clear
        dfvehonboardbefore6 = dfvehonboardbefore6.groupby(['Circle']).size().reset_index(name='Vehicle reported before 6AM to FME')
        dfvehonboardbefore6.loc['Total'] = ['Total', dfvehonboardbefore6['Vehicle reported before 6AM to FME'].sum()]
        vehicle_master = vehicle_master.merge(dfvehonboardbefore6, on="Circle", how="left")
        vehicle_master['Vehicle reported before 6AM to FME'] = vehicle_master['Vehicle reported before 6AM to FME'].fillna(0).astype(int)

        dfvehrelafter10 = daily_vehicle_dump[daily_vehicle_dump['Vehicle release time'] > 22]   #   Need to Clear
        dfvehrelafter10 = dfvehrelafter10.groupby(['Circle']).size().reset_index(name='Vehicle Released after 10PM by FME')
        dfvehrelafter10.loc['Total'] = ['Total', dfvehrelafter10['Vehicle Released after 10PM by FME'].sum()]
        vehicle_master = vehicle_master.merge(dfvehrelafter10, on="Circle", how="left")
        vehicle_master['Vehicle Released after 10PM by FME'] = vehicle_master['Vehicle Released after 10PM by FME'].fillna(0).astype(int)

        dftripnotclosedfme = daily_vehicle_dump[daily_vehicle_dump['Trip Not Closed by'] == "TRIP Not closed by FME"]   #   Need to Clear
        dftripnotclosedfme = dftripnotclosedfme.groupby(['Circle']).size().reset_index(name='Trips not Closed by FME')
        dftripnotclosedfme.loc['Total'] = ['Total', dftripnotclosedfme['Trips not Closed by FME'].sum()]
        vehicle_master = vehicle_master.merge(dftripnotclosedfme, on="Circle", how="left")
        vehicle_master['Trips not Closed by FME'] = vehicle_master['Trips not Closed by FME'].fillna(0).astype(int)

        dftripnotcloseddriver = daily_vehicle_dump[daily_vehicle_dump['Trip Not Closed by'] == "TRIP Not closed by Driver"]   #   Need to Clear
        dftripnotcloseddriver = dftripnotcloseddriver.groupby(['Circle']).size().reset_index(name='Trips not Closed by Drivers')
        dftripnotcloseddriver.loc['Total'] = ['Total', dftripnotcloseddriver['Trips not Closed by Drivers'].sum()]
        vehicle_master = vehicle_master.merge(dftripnotcloseddriver, on="Circle", how="left")
        vehicle_master['Trips not Closed by Drivers'] = vehicle_master['Trips not Closed by Drivers'].fillna(0).astype(int)

        dfvehusedgrth12hr = daily_vehicle_dump[daily_vehicle_dump['Vehicle Trip Time'] == "Used >12 hours"]   #   Need to Clear
        dfvehusedgrth12hr = dfvehusedgrth12hr.groupby(['Circle']).size().reset_index(name='Veh used >12 hour( Only for TRIP closed veh)')
        dfvehusedgrth12hr.loc['Total'] = ['Total', dfvehusedgrth12hr['Veh used >12 hour( Only for TRIP closed veh)'].sum()]
        vehicle_master = vehicle_master.merge(dfvehusedgrth12hr, on="Circle", how="left")
        vehicle_master['Veh used >12 hour( Only for TRIP closed veh)'] = vehicle_master['Veh used >12 hour( Only for TRIP closed veh)'].fillna(0).astype(int)



        vehicle_master['%age of Vehicles used beyond 12 hours'] = (
            (vehicle_master['Veh used >12 hour( Only for TRIP closed veh)'] / vehicle_master['Vehicles Onboarded by FMEs'])
            .fillna(0) * 100
        ).round(0).astype(int).astype(str) + '%'


        vehicle_master['Vehicle Requestd after 10AM %age'] = (
            (vehicle_master['Vehicles Requested after 10AM'] / vehicle_master['Vehicles Requested by FMEs'])
            .fillna(0) * 100
        ).round(0).astype(int).astype(str) + '%'

        vehicle_master['Vehicles onboard after 11AM %age'] = (
            (vehicle_master['Vehicles Onboarded after 11AM'] / vehicle_master['Vehicles Onboarded by FMEs'])
            .fillna(0) * 100
        ).round(0).astype(int).astype(str) + '%'

        vehicle_master['Vehicles not requested %age'] = (
            (vehicle_master['Vehicles Not Requested by FMEs'] / vehicle_master['Total Vehicles in Circle'])
            .fillna(0) * 100
        ).round(0).astype(int).astype(str) + '%'

        vehicle_master['Requested but not accepted %age'] = (
            (vehicle_master['Vehicles requested but not accepted requests'] / vehicle_master['Vehicles Requested by FMEs'])
            .fillna(0) * 100
        ).round(1).astype(str) + '%'

        vehicle_master['Onboard but no wo Handled %age'] = (
            (vehicle_master['Vehicles onboarded but WO not Handled'] / vehicle_master['Vehicles Onboarded by FMEs'])
            .fillna(0) * 100
        ).round(1).astype(str) + '%'

        vehicle_master['Trip not closed %age'] = (
            (vehicle_master['Trips not Closed by FME'] / vehicle_master['Vehicles Onboarded by FMEs'])
            .fillna(0) * 100
        ).round(1).astype(str) + '%'


        #=====================================================================================For Mail Body==========================================================================================================
        # Filter circles where 'Absent Vehicle %age' is greater than 10 and exclude "Total"
        grthtenpercentnotused = vehicle_master.loc[
            (pd.to_numeric(vehicle_master['Absent Vehicle %age'].astype(str).str.rstrip('%'), errors='coerce') > 10) & (vehicle_master['Circle'] != "Total"), 
            'Circle'
        ]
        # Join circle names with commas
        circle_grthtenpercentnotused = ', '.join(grthtenpercentnotused.dropna().astype(str))

        vehusedbeyondtwelvehrs = vehicle_master.loc[
            (pd.to_numeric(vehicle_master['%age of Vehicles used beyond 12 hours'].astype(str).str.rstrip('%'), errors='coerce') > 25) & (vehicle_master['Circle'] != "Total"), 
            'Circle'
        ]
        # Join circle names with commas
        circle_vehusedbeyondtwelvehrs = ', '.join(vehusedbeyondtwelvehrs.dropna().astype(str))

        vehonbbtnowohand = vehicle_master.loc[
            (pd.to_numeric(vehicle_master['Vehicles onboarded but WO not Handled'].astype(str), errors='coerce') > 5) & (vehicle_master['Circle'] != "Total"), 
            'Circle'
        ]
        # Join circle names with commas
        circles_vehonbbtnowohand = ', '.join(vehonbbtnowohand.dropna().astype(str))
      

        #=====================================================================================Styling================================================================================================================

        # Assuming vehicle_master is already defined
        vehicle_master = vehicle_master.style.set_properties(
            subset=pd.IndexSlice[:, :], **{'text-align': 'center', 'white-space': 'normal'}
                )









        # ================ Export Data to Excel ================ #
    # Get yesterday's date in YYYYMMDD format
        yesterday_date = (dt.datetime.now() - dt.timedelta(days=1)).strftime("%Y%m%d")

        # Define header colors
        header_colors = {
            "yellow": "FFFF00",
            "red": "FF0000",
            "light_brown": "FDE9D9",
            "orange": "FFA500"
        }

        # Define header columns for each color
        yellow_headers = ["Circle", "Total Vehicles in Circle", "Vehicles Requested by FMEs",
                        "Vehicles Handled 1 WO", "Vehicles Handled 2-3 WO", "Vehicles Handled >3 WOs","Veh used >12 hour( Only for TRIP closed veh)","%age of Vehicles used beyond 12 hours","Vehicle Requestd after 10AM %age","Vehicles onboard after 11AM %age","Vehicles not requested %age","Requested but not accepted %age"]
        red_headers = ["Vehicles Not Requested by FMEs","Vehicles requested but not accepted requests", "Absent Vehicles","Absent Vehicle %age","Vehicles onboarded but WO not Handled",
                    "Trips not Closed by FME", "Trips not Closed by Drivers","Vehicles Requested after 10AM","Vehicles Onboarded after 11AM",]
        
        red_font_col = ["Vehicles Not Requested by FMEs","Vehicles requested but not accepted requests", "Absent Vehicles","Vehicles onboarded but WO not Handled",
                    "Trips not Closed by FME", "Trips not Closed by Drivers","Vehicles Requested after 10AM","Vehicles Onboarded after 11AM","Veh used >12 hour( Only for TRIP closed veh)"]


        light_brown_headers = ["Total Vehicles Reported ( driver accepted request)",
                                "Vehicles requested but not accepted requests",
                                "Vehicle reported before 6AM to FME", "Vehicle Released after 10PM by FME","Vehicles Onboarded by FMEs","Vehicles handled WOs"]
        orange_headers = ["Onboard but no wo Handled %age","Trip not closed %age"]

        # Define output file path
        output_file_path = os.path.join(output_folder, f"Vehicle_User_Dashboard_{yesterday_date}.xlsx")

        # Step 1: Save Excel using Pandas
        with pd.ExcelWriter(output_file_path, engine='xlsxwriter') as writer:
            vehicle_master.to_excel(writer, sheet_name="Dashboard", index=False)
            daily_vehicle_dump.to_excel(writer, sheet_name="Dump", index=False)

            # Access the worksheet before closing the writer
            worksheet = writer.sheets["Dashboard"]
            worksheet.write(28, 1, ">10% Vehicle not used in:")
            worksheet.write(28, 2, circle_grthtenpercentnotused)

            worksheet.write(29, 1, "Vehicles Onboarded but no WO attached. >5 Veh in :")
            worksheet.write(29, 2, circles_vehonbbtnowohand)

            worksheet.write(30, 1, "Vehicles used >12 hours in circle(>25%):")
            worksheet.write(30, 2, circle_vehusedbeyondtwelvehrs)

        # Step 2: Open saved file and apply formatting using OpenPyXL
        wb = load_workbook(output_file_path)
        ws = wb["Dashboard"]

        # Define white font style
        white_font = Font(color="FFFFFF", bold=True)  # White text with bold formatting

        for cell in ws[1]:  # First row (headers)
            cell_value = str(cell.value).strip()  # Ensure it's a string and remove extra spaces

            if cell_value in yellow_headers:
                cell.fill = PatternFill(start_color=header_colors["yellow"], fill_type="solid")
            elif cell_value in red_headers:
                cell.fill = PatternFill(start_color=header_colors["red"], fill_type="solid")
                cell.font = white_font  # Apply white font ONLY for red headers
            elif cell_value in light_brown_headers:
                cell.fill = PatternFill(start_color=header_colors["light_brown"], fill_type="solid")
            elif cell_value in orange_headers:
                cell.fill = PatternFill(start_color=header_colors["orange"], fill_type="solid")

            # Apply alignment
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


            # Define red font style for values in red_headers columns
            red_font = Font(color="FF0000", bold=False)  # Red text with bold formatting

            # Identify column indexes for red_headers
            red_header_columns = {cell.column for cell in ws[1] if str(cell.value).strip() in red_font_col}

            # Apply red font to all values in identified columns (excluding header)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=min(red_header_columns), max_col=max(red_header_columns)):
                for cell in row:
                    if cell.column in red_header_columns:
                        cell.font = red_font  # Apply red font color



            # Define light brown fill style for values in specific columns
            light_brown_fill = PatternFill(start_color="FDE9D9", fill_type="solid")

            # Identify column indexes for "Circle" and "Total Vehicles in Circle"
            light_brown_columns = {cell.column for cell in ws[1] if str(cell.value).strip() in ["Circle", "Total Vehicles in Circle"]}

            # Excluded row values (text that should not be formatted)
            excluded_rows = {">10% Vehicle not used in:", "Vehicles Onboarded but no WO attached. >5 Veh in :", 
                            "Vehicles used >12 hours in circle(>25%):"}

            # Apply light brown background to non-blank values in identified columns, excluding specific rows
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=min(light_brown_columns), max_col=max(light_brown_columns)):
                row_text = str(row[0].value).strip() if row[0].value else ""  # Assuming first column holds labels
                if row_text not in excluded_rows:  # Exclude specific row labels
                    for cell in row:
                        if cell.column in light_brown_columns and cell.value not in [None, ""]:  # Ignore blank values
                            cell.fill = light_brown_fill  # Apply light brown background




            # List of column headers to process
            percentage_columns = [
                "Absent Vehicle %age",
                "%age of Vehicles used beyond 12 hours",
                "Vehicle Requestd after 10AM %age",
                "Vehicles onboard after 11AM %age",
                "Vehicles not requested %age",
                "Requested but not accepted %age",
                "Onboard but no wo Handled %age",
                "Trip not closed %age"
            ]

            # Dictionary to store column indices
            percentage_col_indices = {}

            # Identify the column indices dynamically
            for col in ws.iter_cols(min_row=1, max_row=1):
                col_header = str(col[0].value).strip() if col[0].value else None
                if col_header in percentage_columns:
                    percentage_col_indices[col_header] = col[0].column

            # Process each identified column
            for col_name, col_index in percentage_col_indices.items():
                # Step 1: Convert % values to decimal (if stored as text, remove % and convert)
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_index, max_col=col_index):
                    for cell in row:
                        if isinstance(cell.value, str) and "%" in cell.value:  # If stored as text
                            try:
                                cell.value = float(cell.value.replace("%", "").strip()) / 100  # Convert to decimal
                            except ValueError:
                                pass  # Ignore non-convertible values
                        elif isinstance(cell.value, (int, float)):  # If numeric
                            cell.value = cell.value / 100 if cell.value > 1 else cell.value

                # Step 2: Apply Conditional Formatting (Red-Yellow-Green)
                color_scale_rule = ColorScaleRule(
                    start_type="num", start_value=0, start_color="00FF00",  # Green for 0%
                    mid_type="percentile", mid_value=50, mid_color="FFFF00",  # Yellow for 50th percentile
                    end_type="num", end_value=1, end_color="FF0000"  # Red for 100%
                )

                # Get column letter
                col_letter = get_column_letter(col_index)
                cell_range = f"{col_letter}2:{col_letter}{ws.max_row}"  # Exclude header row
                ws.conditional_formatting.add(cell_range, color_scale_rule)

                # Step 3: Reapply % format
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_index, max_col=col_index):
                    for cell in row:
                        if isinstance(cell.value, (int, float)):  # Ensure it's numeric
                            cell.number_format = numbers.FORMAT_PERCENTAGE  # Apply % format

                # Define border style (thin border on all sides)
                thin_border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin")
                )

                # Apply border to all non-blank cells
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        if cell.value not in [None, ""]:  # Apply border only to non-blank cells
                            cell.border = thin_border




                # Step 3: Identify and apply bold font to rows containing "Total"
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    if any(cell.value and "total" in str(cell.value).lower() for cell in row):  
                        for cell in row:
                            if cell.value:
                                cell.font = Font(bold=True, color=cell.font.color)  # 


        # Save the updated file
        wb.save(output_file_path)

        # Show success message
        messagebox.showinfo("Success", f"Data saved successfully at:\n{output_file_path}")

    except Exception as e:
        messagebox.showerror("Processing Error", f"An error occurred: {str(e)}")

# ==================================== Tkinter GUI ==================================== #
root = tk.Tk()
root.title("Daily Vehicle Usage Data Processor")
root.geometry("600x150")

# Vehicle Master File Selection
tk.Label(root, text="Vehicle Master File:").grid(row=0, column=0, sticky="w")
vehicle_master_entry = tk.Entry(root, width=50)
vehicle_master_entry.grid(row=0, column=1)
tk.Button(root, text="Browse", command=browse_vehicle_master).grid(row=0, column=2)

# Monthly Dump Folder Selection
tk.Label(root, text="Daily Vehicle Use Dump:").grid(row=1, column=0, sticky="w")
monthly_dump_entry = tk.Entry(root, width=50)
monthly_dump_entry.grid(row=1, column=1)
tk.Button(root, text="Browse", command=browse_monthly_dump).grid(row=1, column=2)

# Output Folder Selection
tk.Label(root, text="Output Folder:").grid(row=2, column=0, sticky="w")
output_folder_entry = tk.Entry(root, width=50)
output_folder_entry.grid(row=2, column=1)
tk.Button(root, text="Browse", command=browse_output_folder).grid(row=2, column=2)

# Process Button
tk.Button(root, text="Generate Report", command=process_files, bg="green", fg="white").grid(row=3, column=1, pady=20)

root.mainloop()
